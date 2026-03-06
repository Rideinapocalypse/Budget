"""
CC Budget & Forecast Tool - Streamlit App
Run with: streamlit run streamlit_app.py
Install:  pip install streamlit openpyxl pandas
"""

import streamlit as st
import pandas as pd
from io import BytesIO
import copy
import urllib.request
import json
import math

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="CC Budget Tool", page_icon="📞", layout="wide",
                   initial_sidebar_state="expanded")

st.markdown("""
<style>
    .block-container { padding-top: 1.5rem; padding-bottom: 2rem; }
    div[data-testid="metric-container"] {
        background: #161b27; border: 1px solid #2a3347;
        border-radius: 8px; padding: 12px 16px;
    }
    div[data-testid="metric-container"] label { color: #5a6480 !important; font-size: 11px !important; }
    div[data-testid="metric-container"] div[data-testid="stMetricValue"] { font-size: 22px !important; }
    .section-title {
        font-size: 11px; font-weight: 700; letter-spacing: 0.1em;
        text-transform: uppercase; color: #5a6480;
        border-bottom: 1px solid #2a3347; padding-bottom: 6px;
        margin-bottom: 12px; margin-top: 8px;
    }
</style>
""", unsafe_allow_html=True)

MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

# ── Multi-client state ───────────────────────────────────────
def _default_client(name="Client A"):
    return dict(
        name=name,
        blocks={m: [] for m in MONTHS},
        cola_configs={},
        overhead_global={
            "TM": {"ratio": 10, "hc_override": None, "salary": 55000},
            "QM": {"ratio": 20, "hc_override": None, "salary": 60000},
            "OM": {"ratio": 50, "hc_override": None, "salary": 80000},
        },
        overhead_monthly={m: None for m in MONTHS},
        opex={
            "training_cost_per_hire": 5000,   # TRY — one-time per backfill hire
            "recruitment_fee":        8000,   # TRY — one-time per any new hire (ramp-up + backfill)
            "it_cost_per_seat":       1500,   # TRY — monthly per active HC
            "facilities_per_seat":    2000,   # TRY — monthly per active HC
            "capex_pc":              15000,   # TRY — one-time per new seat (HC increase only)
            "capex_headset":          3000,   # TRY — one-time per new seat
            "capex_software":         5000,   # TRY — one-time per new seat
        },
        # Actuals: {month: {rev, cost, hc, hrs_billable, margin}}
        actuals={m: {} for m in MONTHS},
    )

if "clients" not in st.session_state:
    st.session_state.clients = [_default_client("Client A")]
for _cl in st.session_state.clients:
    if "actuals" not in _cl:
        _cl["actuals"] = {m: {} for m in MONTHS}
    # Migrate opex fields added in OPEX/CAPEX update
    _opex_defaults = {"training_cost_per_hire":5000,"recruitment_fee":8000,
                      "it_cost_per_seat":1500,"facilities_per_seat":2000,
                      "capex_pc":15000,"capex_headset":3000,"capex_software":5000}
    for _k, _v in _opex_defaults.items():
        _cl.setdefault("opex", {}).setdefault(_k, _v)
# Migrate existing clients that don't have actuals
for _cl in st.session_state.clients:
    if "actuals" not in _cl:
        _cl["actuals"] = {m: {} for m in MONTHS}
    # Migrate opex fields added in OPEX/CAPEX update
    _opex_defaults = {"training_cost_per_hire":5000,"recruitment_fee":8000,
                      "it_cost_per_seat":1500,"facilities_per_seat":2000,
                      "capex_pc":15000,"capex_headset":3000,"capex_software":5000}
    for _k, _v in _opex_defaults.items():
        _cl.setdefault("opex", {}).setdefault(_k, _v)
if "active_client" not in st.session_state:
    st.session_state.active_client = 0
if "active_month" not in st.session_state:
    st.session_state.active_month = "Jan"
if "attrition_rate" not in st.session_state:
    st.session_state.attrition_rate = 0.05
if "backfill_efficiency" not in st.session_state:
    st.session_state.backfill_efficiency = 0.50

def client():
    """Return current active client dict."""
    idx = st.session_state.active_client
    return st.session_state.clients[idx]

def get_oh_cfg(month, cl=None):
    cl = cl or client()
    mo = cl["overhead_monthly"].get(month)
    return mo if mo is not None else cl["overhead_global"]

@st.cache_data(ttl=300)
def fetch_live_fx():
    try:
        url = "https://open.er-api.com/v6/latest/EUR"
        with urllib.request.urlopen(url, timeout=5) as r:
            data = json.loads(r.read())
        return round(data["rates"]["TRY"], 2), True
    except Exception:
        return 38.0, False

def fmt_eur(v): return f"€{v:,.0f}"
def fmt_try(v): return f"₺{v:,.0f}"
def fmt_pct(v): return f"{v*100:.1f}%"

import calendar as _cal
import datetime as _dt

def effective_hc(month, block):
    """Return effective HC for a block in a given month.
    Uses ramp schedule if defined, otherwise falls back to block base HC."""
    ramp = block.get("hc_ramp", {})
    return ramp.get(month, block.get("hc", 0))

def effective_up(month, block_idx, base_up):
    """Return effective unit price for a month, prorated if COLA date falls in it.
    COLA date is treated as a position within the fiscal year (Jan=1 … Dec=12).
    Year is taken from the date only to get days-in-month; comparisons use month number.
    """
    key = str(block_idx)
    cfg = client().get("cola_configs", {}).get(key)
    if not cfg or not cfg.get("date") or not cfg.get("new_up"):
        return base_up
    try:
        cola_date  = _dt.date.fromisoformat(cfg["date"])
        new_up     = float(cfg["new_up"])
        cola_mo    = cola_date.month          # 1–12
        budget_mo  = MONTHS.index(month) + 1  # 1–12

        if budget_mo < cola_mo:
            return base_up        # COLA hasn't happened yet this month
        elif budget_mo > cola_mo:
            return new_up         # COLA fully applied
        else:
            # Transition month — prorate by day
            days_in_mo = _cal.monthrange(cola_date.year, cola_mo)[1]
            days_old   = cola_date.day - 1          # days 1 … (day-1) at old UP
            days_new   = days_in_mo - days_old      # days day … end at new UP
            return (days_old * base_up + days_new * new_up) / days_in_mo
    except Exception:
        return base_up

def get_totals(month, g):
    total_rev_eur = total_cost_eur = total_cost_try = total_rev_try = total_hc = total_hrs = 0.0
    weighted_sal = weighted_fx = weighted_hrs = 0.0
    cl = client()
    for blk_i, b in enumerate(cl["blocks"].get(month, [])):
        raw_shrink = b["shrink_override"] if b.get("shrink_override") is not None else g["shrink"]
        shrink = max(0.0, min(0.99, raw_shrink if raw_shrink <= 1 else raw_shrink / 100))
        fx     = b["fx_override"]     if b.get("fx_override")     is not None else g["fx"]
        hours  = b["hours_override"]  if b.get("hours_override")  is not None else g["hours"]
        hc      = effective_hc(month, b)              # ramp-adjusted HC
        sal     = b.get("salary", 0)
        base_up = b.get("unit_price", 0)
        up      = effective_up(month, blk_i, base_up)  # COLA-adjusted UP
        eff          = hours * (1 - shrink)
        rev_eur      = hc * eff * up
        rev_try      = rev_eur * fx
        cost_try     = hc * sal * g["ctc"] * (1 + g["bonus_pct"]) + hc * g["meal"]
        cost_eur     = cost_try / fx if fx else 0
        total_rev_eur  += rev_eur;  total_rev_try  += rev_try
        total_cost_eur += cost_eur; total_cost_try += cost_try
        total_hc += hc; total_hrs += hc * eff
        weighted_sal += hc * sal
        weighted_fx  += hc * fx
        weighted_hrs += hc * eff

    # Per-block weighted attrition (each block can have own rate, weighted by HC)
    bf_eff = st.session_state.backfill_efficiency
    weighted_att = 0.0
    for b in cl["blocks"].get(month, []):
        hc_b = effective_hc(month, b)
        raw  = b.get("attrition_override")
        rate = raw if raw is not None else st.session_state.attrition_rate
        rate = max(0.0, min(1.0, rate if rate <= 1 else rate / 100))
        weighted_att += hc_b * rate
    att_rate     = (weighted_att / total_hc) if total_hc else st.session_state.attrition_rate
    attrition_hc = weighted_att               # sum of each block's attrition
    backfill_hc  = attrition_hc
    net_hc       = total_hc - attrition_hc

    # Weighted averages for backfill costing
    avg_sal = (weighted_sal / total_hc) if total_hc else 0
    avg_fx  = (weighted_fx  / total_hc) if total_hc else g["fx"]
    avg_eff = (weighted_hrs / total_hc) if total_hc else g["hours"] * (1 - g["shrink"])

    # Backfill cost: full salary (they're employed), partial hours due to training efficiency
    backfill_cost_try  = backfill_hc * avg_sal * g["ctc"] * (1 + g["bonus_pct"]) + backfill_hc * g["meal"]
    backfill_cost_eur  = backfill_cost_try / avg_fx if avg_fx else 0
    # Hours: backfill agents work but at reduced efficiency — counted in produced, not billed
    backfill_hrs       = backfill_hc * avg_eff * bf_eff

    total_cost_try_incl = total_cost_try + backfill_cost_try
    total_cost_eur_incl = total_cost_eur + backfill_cost_eur
    total_hrs_incl      = total_hrs + backfill_hrs

    # OPEX: Training cost (one-time per backfill hire, in TRY)
    opex_cfg = cl.get("opex", {})
    training_cost_try  = backfill_hc * opex_cfg.get("training_cost_per_hire", 0)
    training_cost_eur  = training_cost_try / avg_fx if avg_fx else 0

    # OPEX: Recruitment fee (per new hire = backfill + any HC ramp-up vs prior month)
    prior_month = MONTHS[MONTHS.index(month) - 1] if MONTHS.index(month) > 0 else None
    prior_hc    = sum(effective_hc(prior_month, b) for b in cl["blocks"].get(prior_month, [])) if prior_month else total_hc
    hc_increase = max(0.0, total_hc - prior_hc)   # new seats this month (ramp-up delta)
    new_hires   = backfill_hc + hc_increase        # total new people this month
    recruitment_cost_try = new_hires * opex_cfg.get("recruitment_fee", 0)
    recruitment_cost_eur = recruitment_cost_try / avg_fx if avg_fx else 0

    # OPEX: IT & telephony (monthly × active HC)
    it_cost_try  = total_hc * opex_cfg.get("it_cost_per_seat", 0)
    it_cost_eur  = it_cost_try / avg_fx if avg_fx else 0

    # OPEX: Facilities / rent (monthly × active HC)
    fac_cost_try = total_hc * opex_cfg.get("facilities_per_seat", 0)
    fac_cost_eur = fac_cost_try / avg_fx if avg_fx else 0

    # CAPEX: one-time on HC increase only (new seats, not backfill)
    capex_try = hc_increase * (
        opex_cfg.get("capex_pc", 0) +
        opex_cfg.get("capex_headset", 0) +
        opex_cfg.get("capex_software", 0)
    )
    capex_eur = capex_try / avg_fx if avg_fx else 0

    total_opex_try = training_cost_try + recruitment_cost_try + it_cost_try + fac_cost_try
    total_opex_eur = training_cost_eur + recruitment_cost_eur + it_cost_eur + fac_cost_eur
    total_capex_try = capex_try
    total_capex_eur = capex_eur

    # Overhead roles (TM/QM/OM) — pure cost, no hours, no revenue
    oh = calc_overhead(month, total_hc, g, cl)
    oh_cost_eur = oh["total_cost_eur"]
    oh_cost_try = oh["total_cost_try"]

    grand_cost_eur = total_cost_eur_incl + oh_cost_eur + total_opex_eur + total_capex_eur
    grand_cost_try = total_cost_try_incl + oh_cost_try + total_opex_try + total_capex_try

    # Break-even: must cover all costs (prod + backfill + overhead + opex) per billable hr
    breakeven_up = (grand_cost_eur / total_hrs) if total_hrs > 0 else 0

    return dict(
        rev=total_rev_eur,              rev_try=total_rev_try,
        cost=grand_cost_eur,            cost_try=grand_cost_try,
        cost_excl_backfill=total_cost_eur,
        backfill_cost_eur=backfill_cost_eur,
        backfill_cost_try=backfill_cost_try,
        training_cost_eur=training_cost_eur,
        training_cost_try=training_cost_try,
        recruitment_cost_eur=recruitment_cost_eur,
        recruitment_cost_try=recruitment_cost_try,
        it_cost_eur=it_cost_eur,        it_cost_try=it_cost_try,
        fac_cost_eur=fac_cost_eur,      fac_cost_try=fac_cost_try,
        capex_eur=capex_eur,            capex_try=capex_try,
        hc_increase=hc_increase,        new_hires=new_hires,
        total_opex_eur=total_opex_eur,  total_opex_try=total_opex_try,
        total_capex_eur=total_capex_eur,total_capex_try=total_capex_try,
        oh_cost_eur=oh_cost_eur,        oh_cost_try=oh_cost_try,
        oh=oh,
        margin=total_rev_eur - grand_cost_eur,
        margin_try=total_rev_try - grand_cost_try,
        hc=total_hc,
        hrs=total_hrs_incl,
        hrs_billable=total_hrs,
        backfill_hrs=backfill_hrs,
        attrition_hc=attrition_hc,
        backfill_hc=backfill_hc,
        net_hc=net_hc,
        breakeven_up=breakeven_up,
    )

def calc_overhead(month, prod_hc, g, cl=None):
    """Calculate overhead cost for TM/QM/OM roles for a given month."""
    oh     = get_oh_cfg(month, cl)
    result = {}
    total_cost_try = total_cost_eur = 0.0
    for role, defaults in [("TM",{"ratio":10,"salary":55000}),
                            ("QM",{"ratio":20,"salary":60000}),
                            ("OM",{"ratio":50,"salary":80000})]:
        cfg      = oh.get(role, defaults)
        ratio    = cfg.get("ratio", defaults["ratio"])
        sal      = cfg.get("salary", defaults["salary"])
        override = cfg.get("hc_override")
        # HC: manual override wins, else ratio-based ceiling (you hire whole people)
        if override is not None:
            hc = override
        else:
            hc = math.ceil(prod_hc / ratio) if (ratio > 0 and prod_hc > 0) else 0
        cost_try = hc * sal * g["ctc"] * (1 + g["bonus_pct"]) + hc * g["meal"]
        cost_eur = cost_try / g["fx"] if g["fx"] else 0
        result[role] = dict(hc=hc, salary=sal, ratio=ratio,
                            cost_try=cost_try, cost_eur=cost_eur,
                            manual=override is not None)
        total_cost_try += cost_try
        total_cost_eur += cost_eur
    result["total_cost_try"] = total_cost_try
    result["total_cost_eur"] = total_cost_eur
    return result

# openpyxl helpers
def bdr():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def wcell(ws, r, c, v, bold=False, bg=None, fg="000000", fmt=None, italic=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Calibri", bold=bold, italic=italic, color=fg, size=11)
    cell.border = bdr()
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if bg: cell.fill = PatternFill("solid", start_color=bg)
    if fmt: cell.number_format = fmt
    return cell

def hdr(ws, r, c, v):  return wcell(ws, r, c, v, bold=True,   bg="1F4E79", fg="FFFFFF")
def inp(ws, r, c, v, fmt=None): return wcell(ws, r, c, v, bg="DDEEFF", fg="00008B", fmt=fmt)
def note(ws, r, c, v): return wcell(ws, r, c, v, italic=True,  fg="888888")

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Template builder — 2 sheets only ────────────────────────
def build_template(gh, gs, gfx, gctc, gbp, gm):
    wb = Workbook(); wb.remove(wb.active)

    # ══ Sheet 1: HOW TO USE ══════════════════════════════════
    wi = wb.create_sheet("① How To Use")
    set_widths(wi, [6, 28, 55])
    wi.row_dimensions[1].height = 30

    # Title banner
    tc = wi.cell(row=1, column=1, value="CC Budget Tool — Import Template")
    tc.font   = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    tc.fill   = PatternFill("solid", start_color="1F4E79")
    tc.alignment = Alignment(horizontal="left", vertical="center")
    wi.merge_cells("A1:C1")

    steps = [
        ("①", "Open sheet '② Budget Data'",          "This is the only sheet you need to fill in."),
        ("②", "Fill BLUE cells row by row",           "Each row = one language/team block for that month."),
        ("③", "Month column",                          "Use exactly: Jan  Feb  Mar  Apr  May  Jun  Jul  Aug  Sep  Oct  Nov  Dec"),
        ("④", "Required columns",                      "Month · Language · HC · Base Salary (TRY) · Unit Price (EUR/hr)"),
        ("⑤", "Optional overrides (leave blank)",      "Shrinkage %  ·  FX Rate  ·  Hours/Month  ·  Attrition %  ·  COLA Date  ·  COLA New UP"),
        ("⑥", "COLA date format",                      "YYYY-MM-DD  e.g.  2025-04-15  — the new unit price applies from that date, prorated for the transition month."),
        ("⑦", "Save & import",                         "Save this file → go to the app → sidebar → '⬆ Import Excel'."),
    ]
    for ri, (num, title, detail) in enumerate(steps, start=3):
        wi.row_dimensions[ri].height = 22
        n = wi.cell(row=ri, column=1, value=num)
        n.font = Font(name="Calibri", bold=True, color="1F4E79", size=11)
        t = wi.cell(row=ri, column=2, value=title)
        t.font = Font(name="Calibri", bold=True, color="1F4E79", size=11)
        d = wi.cell(row=ri, column=3, value=detail)
        d.font = Font(name="Calibri", italic=True, color="444444", size=10)

    # Colour legend
    wi.row_dimensions[11].height = 8
    wi.cell(row=12, column=1, value="Colour guide").font = Font(bold=True, color="333333")
    inp(wi, 13, 2, "Blue cell = fill this in")
    note(wi, 13, 3, "White/grey cell = calculated or optional")

    # ══ Sheet 2: BUDGET DATA ══════════════════════════════════
    ws = wb.create_sheet("② Budget Data")
    COLS   = ["Month","Language","HC","Base Salary (TRY)","Unit Price (EUR/hr)",
              "Shrinkage %","FX Rate","Hours/Month","Attrition %","COLA Date","COLA New UP (EUR/hr)"]
    HINTS  = ["Jan … Dec","e.g. DE EN TR","integer","monthly gross","billable €/hr",
              "blank=global","blank=global","blank=global","blank=global","YYYY-MM-DD  blank=none","blank=none"]
    WIDTHS = [10, 18, 8, 20, 20, 14, 12, 14, 13, 18, 22]
    REQUIRED = {0,1,2,3,4}  # Month, Language, HC, Salary, UP — must-fill

    set_widths(ws, WIDTHS)
    ws.freeze_panes = "A5"  # freeze title + header + hints rows

    # Title row
    t = ws.cell(row=1, column=1, value="CC Budget Tool — Budget Data")
    t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", start_color="1F4E79")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    ws.row_dimensions[1].height = 26

    # Sub-title
    st_cell = ws.cell(row=2, column=1,
        value="Fill BLUE cells only. Required: Month, Language, HC, Base Salary, Unit Price. All others are optional overrides.")
    st_cell.font = Font(name="Calibri", italic=True, color="555555", size=10)
    ws.merge_cells(f"A2:{get_column_letter(len(COLS))}2")

    # Header + hint rows
    for ci, (h, hint) in enumerate(zip(COLS, HINTS), 1):
        hdr(ws, 3, ci, h)
        note(ws, 4, ci, hint)
        # mark required columns with a star in header
        if ci-1 in REQUIRED:
            ws.cell(row=3, column=ci).value = h + " *"

    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 14

    # Pre-fill data rows
    ri = 5
    att = st.session_state.attrition_rate
    cola_cfgs = client().get("cola_configs", {})
    for m in MONTHS:
        blocks_m = client()["blocks"].get(m, [])
        rows = blocks_m or [{"lang":"","hc":0,"salary":0,"unit_price":0,
                              "shrink_override":None,"fx_override":None,"hours_override":None}]
        is_first = True
        for blk_i, b in enumerate(rows):
            cola = cola_cfgs.get(str(blk_i), {})
            vals = [
                m if is_first else "",   # month label only on first row of group
                b.get("lang",""),
                b.get("hc", 0),
                b.get("salary", 0),
                b.get("unit_price", 0),
                b["shrink_override"] if b.get("shrink_override") is not None else "",
                b["fx_override"]     if b.get("fx_override")     is not None else "",
                b["hours_override"]  if b.get("hours_override")  is not None else "",
                att,
                cola.get("date",""),
                cola.get("new_up",""),
            ]
            for ci, v in enumerate(vals, 1):
                # required cols → blue input style; optional → lighter
                if ci-1 in REQUIRED:
                    inp(ws, ri, ci, v)
                else:
                    c = ws.cell(row=ri, column=ci, value=v)
                    c.font   = Font(name="Calibri", color="333333", size=11)
                    c.fill   = PatternFill("solid", start_color="F0F4FF")
                    c.border = bdr()
                    c.alignment = Alignment(horizontal="left", vertical="center")
            is_first = False
            ri += 1

        # Light divider row between month groups
        for ci in range(1, len(COLS)+1):
            c = ws.cell(row=ri, column=ci, value="")
            c.fill = PatternFill("solid", start_color="E8EDF5")
            c.border = bdr()
        ri += 1

    buf = BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

# ── Export builder — 3 clean sheets ─────────────────────────
def build_export(g):
    wb = Workbook(); wb.remove(wb.active)

    # helper — number cell
    def num(ws, r, c, v, fmt="#,##0", bold=False, bg=None):
        cell = wcell(ws, r, c, v, bold=bold, bg=bg)
        cell.number_format = fmt
        return cell

    # ══ Sheet 1: P&L SUMMARY ═════════════════════════════════
    ws1 = wb.create_sheet("P&L Summary")
    PL_COLS = ["Month","Revenue (EUR)","Prod Cost (EUR)","Backfill Cost (EUR)",
               "Overhead Cost (EUR)","Total Cost (EUR)","Gross Margin (EUR)",
               "Margin %","Break-even €/hr","Avg Selling €/hr",
               "Prod HC","TM HC","QM HC","OM HC","Net HC (EOM)"]
    set_widths(ws1, [10,16,16,16,16,16,16,10,14,14,10,8,8,8,14])

    # title
    t = ws1.cell(row=1, column=1, value="CC Budget — P&L Summary")
    t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", start_color="1F4E79")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws1.merge_cells(f"A1:{get_column_letter(len(PL_COLS))}1")
    ws1.row_dimensions[1].height = 26

    for ci, h in enumerate(PL_COLS, 1): hdr(ws1, 2, ci, h)

    fy = {k: 0.0 for k in ["rev","cost","margin","cost_excl_backfill",
                             "backfill_cost_eur","oh_cost_eur","hrs_billable"]}
    for ri, m in enumerate(MONTHS, start=3):
        t_m = get_totals(m, g)
        oh  = t_m["oh"]
        avg_up = t_m["rev"] / t_m["hrs_billable"] if t_m["hrs_billable"] else 0
        row_vals = [
            m,
            t_m["rev"], t_m["cost_excl_backfill"], t_m["backfill_cost_eur"],
            t_m["oh_cost_eur"], t_m["cost"], t_m["margin"],
            t_m["margin"]/t_m["rev"] if t_m["rev"] else 0,
            t_m["breakeven_up"], avg_up,
            t_m["hc"], oh["TM"]["hc"], oh["QM"]["hc"], oh["OM"]["hc"], t_m["net_hc"],
        ]
        for ci, v in enumerate(row_vals, 1):
            if ci == 1:
                wcell(ws1, ri, ci, v, bold=True)
            elif ci == 8:
                num(ws1, ri, ci, v, fmt="0.0%")
            elif ci in (9,10):
                num(ws1, ri, ci, v, fmt='€#,##0.00"/hr"')
            elif ci > 10:
                num(ws1, ri, ci, v, fmt="0.00")
            else:
                num(ws1, ri, ci, v)
        # accumulate FY
        for k in fy:
            fy[k] += t_m.get(k, 0)

    # Full Year row
    fy_ri = len(MONTHS) + 3
    fy_avg = fy["rev"] / fy["hrs_billable"] if fy["hrs_billable"] else 0
    fy_be  = fy["cost"] / fy["hrs_billable"] if fy["hrs_billable"] else 0
    fy_row = [
        "Full Year",
        fy["rev"], fy["cost_excl_backfill"], fy["backfill_cost_eur"],
        fy["oh_cost_eur"], fy["cost"], fy["margin"],
        fy["margin"]/fy["rev"] if fy["rev"] else 0,
        fy_be, fy_avg,
        "","","","","",
    ]
    for ci, v in enumerate(fy_row, 1):
        bg = "E8F0FE"
        if ci == 1: wcell(ws1, fy_ri, ci, v, bold=True, bg=bg)
        elif ci == 8: num(ws1, fy_ri, ci, v, fmt="0.0%", bold=True, bg=bg)
        elif ci in (9,10): num(ws1, fy_ri, ci, v, fmt='€#,##0.00"/hr"', bold=True, bg=bg)
        elif isinstance(v, float): num(ws1, fy_ri, ci, v, bold=True, bg=bg)
        else: wcell(ws1, fy_ri, ci, v, bold=True, bg=bg)

    # ══ Sheet 2: BLOCK DETAIL ════════════════════════════════
    ws2 = wb.create_sheet("Block Detail")
    BD_COLS = ["Month","Block","Language","HC","Base Salary (TRY)","Unit Price (EUR/hr)",
               "Eff. UP (EUR/hr)","Eff. Hours/Agent","Revenue (EUR)","Prod Cost (EUR)","Margin (EUR)","Margin %"]
    set_widths(ws2, [10,8,16,8,18,18,16,16,16,16,16,10])

    t2 = ws2.cell(row=1, column=1, value="CC Budget — Block Detail")
    t2.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t2.fill = PatternFill("solid", start_color="1F4E79")
    t2.alignment = Alignment(horizontal="left", vertical="center")
    ws2.merge_cells(f"A1:{get_column_letter(len(BD_COLS))}1")
    ws2.row_dimensions[1].height = 26
    for ci, h in enumerate(BD_COLS, 1): hdr(ws2, 2, ci, h)

    ri2 = 3
    for m in MONTHS:
        for blk_i, b in enumerate(client()["blocks"].get(m, [])):
            raw_shrink = b["shrink_override"] if b.get("shrink_override") is not None else g["shrink"]
            shrink = max(0.0, min(0.99, raw_shrink if raw_shrink <= 1 else raw_shrink / 100))
            fx     = b["fx_override"]    if b.get("fx_override")    is not None else g["fx"]
            hours  = b["hours_override"] if b.get("hours_override") is not None else g["hours"]
            hc, sal = b.get("hc",0), b.get("salary",0)
            base_up = b.get("unit_price",0)
            eff_up  = effective_up(m, blk_i, base_up)
            eff_hrs = hours * (1 - shrink)
            rev     = hc * eff_hrs * eff_up
            cost    = (hc * sal * g["ctc"] * (1 + g["bonus_pct"]) + hc * g["meal"]) / fx if fx else 0
            margin  = rev - cost
            row_vals = [m, f"#{blk_i+1}", b.get("lang",""), hc, sal, base_up,
                        eff_up, eff_hrs, rev, cost, margin,
                        margin/rev if rev else 0]
            for ci, v in enumerate(row_vals, 1):
                if ci in (1,2,3): wcell(ws2, ri2, ci, v)
                elif ci == 12:    num(ws2, ri2, ci, v, fmt="0.0%")
                elif ci in (5,):  num(ws2, ri2, ci, v, fmt="#,##0")
                elif ci in (6,7,8): num(ws2, ri2, ci, v, fmt="#,##0.00")
                else:             num(ws2, ri2, ci, v)
            ri2 += 1

    # ══ Sheet 3: SETTINGS SNAPSHOT ═══════════════════════════
    ws3 = wb.create_sheet("Settings Snapshot")
    set_widths(ws3, [32, 20])
    t3 = ws3.cell(row=1, column=1, value="Global Settings at time of export")
    t3.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    t3.fill = PatternFill("solid", start_color="1F4E79")
    t3.alignment = Alignment(horizontal="left", vertical="center")
    ws3.merge_cells("A1:B1")
    ws3.row_dimensions[1].height = 24

    hdr(ws3,2,1,"Setting"); hdr(ws3,2,2,"Value")
    settings = [
        ("Worked Hours / Agent / Month", g["hours"]),
        ("Shrinkage % (global)",         f"{g['shrink']*100:.1f}%"),
        ("FX Rate (1 EUR = TRY)",        g["fx"]),
        ("CTC Multiplier",               g["ctc"]),
        ("Bonus % of Base",              f"{g['bonus_pct']*100:.1f}%"),
        ("Meal Card / Agent / Month (TRY)", g["meal"]),
        ("Monthly Attrition Rate",       f"{st.session_state.attrition_rate*100:.1f}%"),
        ("Backfill Training Efficiency", f"{st.session_state.backfill_efficiency*100:.0f}%"),
        ("Export date",                  _dt.date.today().isoformat()),
    ]
    for ri3, (lbl, val) in enumerate(settings, start=3):
        wcell(ws3, ri3, 1, lbl)
        wcell(ws3, ri3, 2, val)

    buf = BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

# ── PDF Report builder ───────────────────────────────────────
def build_pdf(g):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    # Colours
    DARK   = colors.HexColor("#0e1420")
    NAVY   = colors.HexColor("#1F4E79")
    BLUE   = colors.HexColor("#3b82f6")
    GREEN  = colors.HexColor("#10b981")
    RED    = colors.HexColor("#ef4444")
    AMBER  = colors.HexColor("#f59e0b")
    LIGHT  = colors.HexColor("#e8edf5")
    MID    = colors.HexColor("#8b96b0")
    ROW_A  = colors.HexColor("#131929")
    ROW_B  = colors.HexColor("#0e1420")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", fontSize=20, textColor=LIGHT,
                                  fontName="Helvetica-Bold", spaceAfter=4)
    sub_style   = ParagraphStyle("sub",   fontSize=10, textColor=MID,
                                  fontName="Helvetica", spaceAfter=16)
    h2_style    = ParagraphStyle("h2",    fontSize=13, textColor=LIGHT,
                                  fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=6)
    note_style  = ParagraphStyle("note",  fontSize=8,  textColor=MID,
                                  fontName="Helvetica-Oblique")

    def p(text, style=None): return Paragraph(text, style or styles["Normal"])

    story = []

    # ── Cover block ──────────────────────────────────────────
    story.append(Paragraph(f"CC Budget Report", title_style))
    story.append(Paragraph(f"Client: <b>{client()['name']}</b>  ·  Generated: {_dt.date.today().isoformat()}", sub_style))
    story.append(HRFlowable(width="100%", color=NAVY, thickness=1.5, spaceAfter=14))

    # ── KPI summary row ──────────────────────────────────────
    all_totals = [get_totals(m, g) for m in MONTHS]
    fy_rev    = sum(t["rev"]    for t in all_totals)
    fy_cost   = sum(t["cost"]   for t in all_totals)
    fy_margin = sum(t["margin"] for t in all_totals)
    fy_mgn_pct= fy_margin/fy_rev*100 if fy_rev else 0
    fy_hc     = max(t["hc"] for t in all_totals)

    kpi_data = [
        ["Full Year Revenue", "Total Cost", "Gross Margin", "Margin %", "Peak HC"],
        [f"€{fy_rev:,.0f}", f"€{fy_cost:,.0f}", f"€{fy_margin:,.0f}", f"{fy_mgn_pct:.1f}%", f"{int(fy_hc)}"],
    ]
    kpi_col_w = [3.2*cm]*5
    kpi_tbl = Table(kpi_data, colWidths=kpi_col_w)
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), NAVY),
        ("TEXTCOLOR",  (0,0), (-1,0), LIGHT),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,0), 8),
        ("BACKGROUND", (0,1), (-1,1), ROW_A),
        ("TEXTCOLOR",  (0,1), (-1,1), LIGHT),
        ("FONTNAME",   (0,1), (-1,1), "Helvetica-Bold"),
        ("FONTSIZE",   (0,1), (-1,1), 11),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1),(-1,1), [ROW_A]),
        ("GRID",       (0,0), (-1,-1), 0.4, colors.HexColor("#2a3347")),
        ("TOPPADDING", (0,0), (-1,-1), 7),
        ("BOTTOMPADDING", (0,0), (-1,-1), 7),
    ]))
    story.append(kpi_tbl)
    story.append(Spacer(1, 14))

    # ── Monthly P&L table ────────────────────────────────────
    story.append(Paragraph("Monthly P&L", h2_style))
    story.append(HRFlowable(width="100%", color=NAVY, thickness=0.5, spaceAfter=6))

    PL_HDR = ["Month","Revenue","Prod Cost","Backfill","Training","Overhead","Total Cost","GM","GM%","BE €/hr"]
    pl_rows = [PL_HDR]
    for m, t in zip(MONTHS, all_totals):
        oh_tot = t["oh_cost_eur"]
        avg_up = t["rev"]/t["hrs_billable"] if t["hrs_billable"] else 0
        mg_col = GREEN if t["margin"] >= 0 else RED
        pl_rows.append([
            m,
            f"€{t['rev']:,.0f}",
            f"€{t['cost_excl_backfill']:,.0f}",
            f"€{t['backfill_cost_eur']:,.0f}",
            f"€{t.get('training_cost_eur',0):,.0f}",
            f"€{oh_tot:,.0f}",
            f"€{t['cost']:,.0f}",
            f"€{t['margin']:,.0f}",
            f"{t['margin']/t['rev']*100:.1f}%" if t['rev'] else "—",
            f"€{t['breakeven_up']:.2f}",
        ])
    # Full Year row
    fy_oh = sum(t["oh_cost_eur"] for t in all_totals)
    fy_bf = sum(t["backfill_cost_eur"] for t in all_totals)
    fy_tr = sum(t.get("training_cost_eur",0) for t in all_totals)
    fy_pc = sum(t["cost_excl_backfill"] for t in all_totals)
    fy_be = fy_cost / sum(t["hrs_billable"] for t in all_totals) if sum(t["hrs_billable"] for t in all_totals) else 0
    pl_rows.append([
        "Full Year",
        f"€{fy_rev:,.0f}", f"€{fy_pc:,.0f}", f"€{fy_bf:,.0f}", f"€{fy_tr:,.0f}",
        f"€{fy_oh:,.0f}", f"€{fy_cost:,.0f}", f"€{fy_margin:,.0f}",
        f"{fy_mgn_pct:.1f}%", f"€{fy_be:.2f}",
    ])

    col_w = [1.4*cm, 2.0*cm, 2.0*cm, 1.8*cm, 1.8*cm, 1.8*cm, 2.0*cm, 2.0*cm, 1.3*cm, 1.8*cm]
    pl_tbl = Table(pl_rows, colWidths=col_w, repeatRows=1)
    ts = [
        ("BACKGROUND",    (0,0),  (-1,0),  NAVY),
        ("TEXTCOLOR",     (0,0),  (-1,0),  LIGHT),
        ("FONTNAME",      (0,0),  (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),  (-1,-1), 7.5),
        ("ALIGN",         (1,0),  (-1,-1), "RIGHT"),
        ("ALIGN",         (0,0),  (0,-1),  "LEFT"),
        ("GRID",          (0,0),  (-1,-1), 0.3, colors.HexColor("#2a3347")),
        ("TOPPADDING",    (0,0),  (-1,-1), 4),
        ("BOTTOMPADDING", (0,0),  (-1,-1), 4),
        # Full Year row bold
        ("FONTNAME",  (0,len(pl_rows)-1), (-1,len(pl_rows)-1), "Helvetica-Bold"),
        ("BACKGROUND",(0,len(pl_rows)-1), (-1,len(pl_rows)-1), colors.HexColor("#1a2540")),
        ("TEXTCOLOR", (0,len(pl_rows)-1), (-1,len(pl_rows)-1), LIGHT),
    ]
    # Alternating rows
    for r in range(1, len(pl_rows)-1):
        bg = ROW_A if r % 2 == 1 else ROW_B
        ts.append(("BACKGROUND", (0,r), (-1,r), bg))
        ts.append(("TEXTCOLOR",  (0,r), (-1,r), LIGHT))
        # Colour margin cell
        margin_val = all_totals[r-1]["margin"]
        mg_c = colors.HexColor("#10b981") if margin_val >= 0 else colors.HexColor("#ef4444")
        ts.append(("TEXTCOLOR", (7,r), (7,r), mg_c))
        ts.append(("TEXTCOLOR", (8,r), (8,r), mg_c))
    pl_tbl.setStyle(TableStyle(ts))
    story.append(pl_tbl)

    # ── Actuals vs Budget table (if any actuals entered) ────
    has_act = any(bool(client()["actuals"].get(m,{}).get("rev") or
                       client()["actuals"].get(m,{}).get("cost")) for m in MONTHS)
    if has_act:
        story.append(Paragraph("Actual vs Budget", h2_style))
        story.append(HRFlowable(width="100%", color=NAVY, thickness=0.5, spaceAfter=6))
        avb_hdr = ["Month","Bgt Rev","Act Rev","Rev Var","Bgt GM","Act GM","GM Var"]
        avb_pdf = [avb_hdr]
        for m in MONTHS:
            mt  = all_totals[MONTHS.index(m)]
            act = client()["actuals"].get(m, {})
            if not (act.get("rev") or act.get("cost")): continue
            a_rev = act.get("rev",0); a_gm = act.get("margin", a_rev - act.get("cost",0))
            rv = a_rev - mt["rev"]; gv = a_gm - mt["margin"]
            avb_pdf.append([
                m,
                f"€{mt['rev']:,.0f}", f"€{a_rev:,.0f}",
                f"{'+' if rv>=0 else ''}€{rv:,.0f}",
                f"€{mt['margin']:,.0f}", f"€{a_gm:,.0f}",
                f"{'+' if gv>=0 else ''}€{gv:,.0f}",
            ])
        avb_tbl = Table(avb_pdf, colWidths=[1.5*cm,2.2*cm,2.2*cm,2.2*cm,2.2*cm,2.2*cm,2.2*cm], repeatRows=1)
        avb_ts = [
            ("BACKGROUND",(0,0),(-1,0), NAVY), ("TEXTCOLOR",(0,0),(-1,0), LIGHT),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"), ("FONTSIZE",(0,0),(-1,-1), 7.5),
            ("ALIGN",(1,0),(-1,-1),"RIGHT"), ("ALIGN",(0,0),(0,-1),"LEFT"),
            ("GRID",(0,0),(-1,-1), 0.3, colors.HexColor("#2a3347")),
            ("TOPPADDING",(0,0),(-1,-1), 4), ("BOTTOMPADDING",(0,0),(-1,-1), 4),
        ]
        for r in range(1, len(avb_pdf)):
            bg = ROW_A if r % 2 == 1 else ROW_B
            avb_ts += [("BACKGROUND",(0,r),(-1,r), bg), ("TEXTCOLOR",(0,r),(-1,r), LIGHT)]
            rv_val = avb_pdf[r][3]
            gv_val = avb_pdf[r][6]
            avb_ts.append(("TEXTCOLOR",(3,r),(3,r), GREEN if rv_val.startswith("+") else RED))
            avb_ts.append(("TEXTCOLOR",(6,r),(6,r), GREEN if gv_val.startswith("+") else RED))
        avb_tbl.setStyle(TableStyle(avb_ts))
        story.append(avb_tbl)
        story.append(Spacer(1, 14))

    # ── Footer ───────────────────────────────────────────────
    story.append(Spacer(1, 18))
    story.append(HRFlowable(width="100%", color=NAVY, thickness=0.5))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f"CC Budget Tool  ·  FX: 1 EUR = ₺{g['fx']:,.2f}  ·  CTC: {g['ctc']}x  ·  "
        f"Attrition: {st.session_state.attrition_rate*100:.1f}%  ·  "
        f"Training cost/hire: ₺{client().get('opex',{}).get('training_cost_per_hire',0):,.0f}",
        note_style))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()

# ── SIDEBAR ───────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 📞 CCBudget")
    st.caption("Call Center Forecast Tool")
    st.divider()

    st.markdown('<div class="section-title">Global Inputs</div>', unsafe_allow_html=True)
    g_hours  = st.number_input("Worked Hours / Agent / Month", value=180, step=1, min_value=1)
    g_shrink = st.slider("Shrinkage % (default)", 0.0, 0.5, 0.15, 0.01, format="%.0f%%")

    live_fx, fx_ok = fetch_live_fx()
    if fx_ok:
        st.caption(f"🟢 Live EUR/TRY: **{live_fx}** (auto-fetched, editable below)")
    else:
        st.caption("🔴 Could not fetch live rate — using fallback")
    g_fx = st.number_input("FX Rate (1 EUR = TRY)", value=live_fx, step=0.5, min_value=0.1)

    st.divider()
    st.markdown('<div class="section-title">Global Cost Drivers</div>', unsafe_allow_html=True)
    g_ctc       = st.number_input("Salary Multiplier (CTC)", value=1.70, step=0.05, min_value=1.0)
    g_bonus_pct = st.number_input("Bonus % of Base Salary",  value=0.10, step=0.01, min_value=0.0)
    g_meal      = st.number_input("Meal Card / Agent / Month (TRY)", value=5850, step=50, min_value=0)

    st.divider()
    st.markdown('<div class="section-title">Attrition & Backfill</div>', unsafe_allow_html=True)
    attrition_pct = st.slider("Monthly Attrition %", 0.0, 0.30, 0.05, 0.005, format="%.1f%%",
                               help="Fraction of HC lost per month. Backfill hired 1-for-1.")
    bf_efficiency = st.slider("Backfill Training Efficiency %", 0.0, 1.0, 0.50, 0.05, format="%.0f%%",
                               help="How productive backfill agents are while in training. 50% = half speed. Hours are counted but generate no revenue.")
    st.session_state.attrition_rate       = attrition_pct
    st.session_state.backfill_efficiency  = bf_efficiency

    st.divider()
    st.markdown('<div class="section-title">OPEX & CAPEX</div>', unsafe_allow_html=True)
    st.caption("All rates in TRY. Converted to EUR at weighted FX rate.")

    with st.expander("🔁 OPEX — Recurring (per HC / month)", expanded=False):
        opex_it  = st.number_input("IT & Telephony per seat / month (TRY)",
            value=float(client().get("opex",{}).get("it_cost_per_seat",1500)),
            step=100.0, min_value=0.0,
            help="Softphone license, headset rental, CRM seat, internet share.")
        opex_fac = st.number_input("Facilities / Rent per seat / month (TRY)",
            value=float(client().get("opex",{}).get("facilities_per_seat",2000)),
            step=100.0, min_value=0.0,
            help="Desk cost: office rent, electricity, cleaning allocated per seat.")
        client().setdefault("opex",{}).update({
            "it_cost_per_seat": opex_it, "facilities_per_seat": opex_fac})

    with st.expander("🧾 OPEX — One-time per new hire", expanded=False):
        opex_training = st.number_input("Training cost per backfill hire (TRY)",
            value=float(client().get("opex",{}).get("training_cost_per_hire",5000)),
            step=500.0, min_value=0.0,
            help="Triggered by attrition backfill only. Onboarding, materials, trainer time.")
        opex_recruit  = st.number_input("Recruitment fee per new hire (TRY)",
            value=float(client().get("opex",{}).get("recruitment_fee",8000)),
            step=500.0, min_value=0.0,
            help="Triggered by BOTH backfill AND ramp-up HC increases. Agency or internal HR cost.")
        client().setdefault("opex",{}).update({
            "training_cost_per_hire": opex_training, "recruitment_fee": opex_recruit})

    with st.expander("🖥 CAPEX — One-time per new seat (ramp-up only)", expanded=False):
        st.caption("Only charged when HC increases vs prior month. Not charged on backfill or stable months.")
        capex_pc  = st.number_input("Workstation / PC (TRY)",
            value=float(client().get("opex",{}).get("capex_pc",15000)),
            step=500.0, min_value=0.0)
        capex_hs  = st.number_input("Headset hardware (TRY)",
            value=float(client().get("opex",{}).get("capex_headset",3000)),
            step=100.0, min_value=0.0)
        capex_sw  = st.number_input("Software license / perpetual (TRY)",
            value=float(client().get("opex",{}).get("capex_software",5000)),
            step=500.0, min_value=0.0)
        total_per_seat = capex_pc + capex_hs + capex_sw
        st.markdown(
            f"<div style='background:#1e2535;border:1px solid #2a3347;border-radius:5px;"
            f"padding:8px 12px;font-size:12px;color:#8b96b0'>"
            f"Total per new seat: <b style='color:#e8edf5'>₺{total_per_seat:,.0f}</b>"
            f"</div>", unsafe_allow_html=True)
        client().setdefault("opex",{}).update({
            "capex_pc": capex_pc, "capex_headset": capex_hs, "capex_software": capex_sw})

    g = dict(hours=g_hours, shrink=g_shrink, fx=g_fx,
             ctc=g_ctc, bonus_pct=g_bonus_pct, meal=g_meal)

    st.divider()
    st.markdown('<div class="section-title">Data Import / Export</div>', unsafe_allow_html=True)

    st.download_button(
        label="📋 Download Blank Template",
        data=build_template(g_hours, g_shrink, g_fx, g_ctc, g_bonus_pct, g_meal),
        file_name="CC_Budget_Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        help="Fillable Excel template — fill blue cells, then import.",
    )
    st.download_button(
        label="⬇ Export Data to Excel",
        data=build_export(g),
        file_name="CC_Budget_Export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )
    try:
        pdf_data = build_pdf(g)
        st.download_button(
            label="📄 Export PDF Report",
            data=pdf_data,
            file_name=f"CC_Budget_{client()['name'].replace(' ','_')}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
    except Exception as pdf_err:
        st.caption(f"PDF unavailable: {pdf_err}")

    st.divider()
    uploaded = st.file_uploader("⬆ Import Excel", type=["xlsx"])
    if uploaded:
        try:
            xls = pd.ExcelFile(uploaded)
            def sf(v):
                try: return float(v) if str(v).strip() not in ("","nan") else None
                except: return None

            loaded = 0
            # Support both old multi-sheet format and new single-sheet format
            if "Budget Data" in xls.sheet_names:
                # New single-sheet format
                df_all = pd.read_excel(xls, sheet_name="Budget Data", header=2)
                df_all.columns = [str(c).strip() for c in df_all.columns]
                df_all = df_all[pd.to_numeric(df_all.get("HC", pd.Series()), errors="coerce").notna()]
                new_blocks = {m: [] for m in MONTHS}
                for _, row in df_all.iterrows():
                    m = str(row.get("Month","")).strip()
                    if m not in MONTHS: continue
                    new_blocks[m].append({
                        "lang":            str(row.get("Language","")).strip() if pd.notna(row.get("Language","")) else "",
                        "hc":              int(float(row["HC"])) if pd.notna(row.get("HC")) else 0,
                        "salary":          float(row.get("Base Salary (TRY)",0)) if pd.notna(row.get("Base Salary (TRY)")) else 0,
                        "unit_price":      float(row.get("Unit Price (EUR/hr)",0)) if pd.notna(row.get("Unit Price (EUR/hr)")) else 0,
                        "shrink_override": sf(row.get("Shrinkage Override","")),
                        "fx_override":     sf(row.get("FX Override","")),
                        "hours_override":  sf(row.get("Hours Override","")),
                    })
                    loaded += 1
                client()["blocks"] = new_blocks
            else:
                # Legacy per-month sheet format
                for m in MONTHS:
                    if m not in xls.sheet_names: continue
                    df_m = pd.read_excel(xls, sheet_name=m)
                    if "HC" in df_m.columns:
                        df_m = df_m[pd.to_numeric(df_m["HC"], errors="coerce").notna()]
                    blocks = []
                    for _, row in df_m.iterrows():
                        blocks.append({
                            "lang":            str(row.get("Language","")).strip() if pd.notna(row.get("Language","")) else "",
                            "hc":              int(float(row["HC"])) if pd.notna(row.get("HC")) else 0,
                            "salary":          float(row.get("Base Salary (TRY)",0)) if pd.notna(row.get("Base Salary (TRY)")) else 0,
                            "unit_price":      float(row.get("Unit Price (EUR/hr)",0)) if pd.notna(row.get("Unit Price (EUR/hr)")) else 0,
                            "shrink_override": sf(row.get("Shrinkage Override","")),
                            "fx_override":     sf(row.get("FX Override","")),
                            "hours_override":  sf(row.get("Hours Override","")),
                        })
                    client()["blocks"][m] = blocks
                    loaded += len(blocks)
            st.success(f"✅ Imported {loaded} blocks across all months!")
            st.rerun()
        except Exception as e:
            st.error(f"Import failed: {e}")

# ── MAIN ──────────────────────────────────────────────────────
st.markdown("## 📞 CC Budget & Forecast")

# ── Client tabs ───────────────────────────────────────────────
cl_row = st.columns([8, 2])
with cl_row[0]:
    client_names = [c["name"] for c in st.session_state.clients]
    # Render as buttons
    btn_cols = st.columns(min(len(client_names) + 1, 10))
    for ci, cname in enumerate(client_names):
        is_active = ci == st.session_state.active_client
        if btn_cols[ci].button(
            f"{'✦ ' if is_active else ''}{cname}",
            key=f"cl_tab_{ci}",
            type="primary" if is_active else "secondary",
            use_container_width=True,
        ):
            st.session_state.active_client = ci
            st.rerun()
    if len(client_names) < 8:
        if btn_cols[len(client_names)].button("＋ Add Client", key="add_cl", use_container_width=True):
            n = len(st.session_state.clients) + 1
            st.session_state.clients.append(_default_client(f"Client {chr(64+n)}"))
            st.session_state.active_client = n - 1
            st.rerun()

with cl_row[1]:
    cl_name_input = st.text_input("Rename client", value=client()["name"],
                                   key=f"cl_name_{st.session_state.active_client}",
                                   label_visibility="collapsed", placeholder="Client name")
    if cl_name_input != client()["name"]:
        client()["name"] = cl_name_input

    if len(st.session_state.clients) > 1:
        if st.button("🗑 Remove this client", key="del_cl", use_container_width=True):
            st.session_state.clients.pop(st.session_state.active_client)
            st.session_state.active_client = max(0, st.session_state.active_client - 1)
            st.rerun()

st.markdown("---")

cols_tabs = st.columns(12)
for i, m in enumerate(MONTHS):
    with cols_tabs[i]:
        if st.button(m, use_container_width=True,
                     type="primary" if m == st.session_state.active_month else "secondary"):
            st.session_state.active_month = m
            st.rerun()

active = st.session_state.active_month
st.markdown(f"### {active}")

t = get_totals(active, g)
k1,k2,k3,k4,k5 = st.columns(5)
k1.metric("Revenue (EUR)", fmt_eur(t["rev"]),
          delta=fmt_try(t["rev_try"]) + " TRY", delta_color="off")
k2.metric("Cost (EUR)", fmt_eur(t["cost"]),
          delta=f'{fmt_try(t["cost_try"])} TRY  |  OH: €{t["oh_cost_eur"]:,.0f}',
          delta_color="off", help="Includes production, backfill & overhead (TM/QM/OM)")
gm_pct = fmt_pct(t["margin"]/t["rev"]) if t["rev"] else "0%"
k3.metric(f"Gross Margin (EUR)  {gm_pct}", fmt_eur(t["margin"]),
          delta=f'{fmt_try(t["margin_try"])} TRY  |  {gm_pct}',
          delta_color="normal")
attr_hc   = t["attrition_hc"]
net_hc    = t["net_hc"]
attr_pct  = fmt_pct(attrition_pct)
k4.metric("Total HC",
          f"{int(t['hc'])} agents",
          delta=f"-{attr_hc} attrition ({attr_pct})",
          delta_color="inverse",
          help=f"Net HC after {attr_pct} attrition: {net_hc} agents")
k5.metric("Produced Hrs",
          f"{t['hrs']:,.0f} hrs",
          delta=f"Billable: {t['hrs_billable']:,.0f}  |  Backfill: {t['backfill_hrs']:,.0f}",
          delta_color="off",
          help="Total hours worked incl. backfill trainees. Only billable hours generate revenue.")

# Break-even banner
if t["hc"] > 0:
    be   = t["breakeven_up"]
    avg_up = t["rev"] / t["hrs_billable"] if t["hrs_billable"] else 0
    be_gap = avg_up - be
    be_color = "#10b981" if be_gap >= 0 else "#ef4444"
    be_icon  = "✅" if be_gap >= 0 else "⚠️"
    st.markdown(
        f"<div style='background:#12192a;border:1px solid #2a3347;border-radius:6px;"
        f"padding:8px 18px;margin-top:-8px;margin-bottom:4px;"
        f"display:flex;justify-content:space-between;align-items:center;font-size:13px'>"
        f"<span style='color:#5a6480'>Break-even Unit Price</span>"
        f"<span style='color:{be_color};font-weight:700'>{be_icon} €{be:.2f}/hr break-even"
        f"&nbsp;·&nbsp;avg selling €{avg_up:.2f}/hr"
        f"&nbsp;·&nbsp;gap <b>€{be_gap:+.2f}/hr</b></span>"
        f"</div>", unsafe_allow_html=True
    )

# ── Attrition warning banner ─────────────────────────────────
if t["hc"] > 0:
    attr_hc  = t["attrition_hc"]
    net_hc   = t["net_hc"]
    color    = "#f59e0b" if attr_hc > 0 else "#10b981"
    icon     = "⚠️" if attr_hc > 0 else "✅"
    st.markdown(
        f"<div style='background:#1e2535;border:1px solid {color};border-radius:6px;"
        f"padding:10px 18px;margin-bottom:12px;display:flex;justify-content:space-between;align-items:center'>"
        f"<span style='color:{color};font-weight:600'>{icon} Attrition Forecast — {active}</span>"
        f"<span style='color:#e8edf5'>"
        f"Starting HC: <b>{int(t['hc'])}</b> &nbsp;|&nbsp; "
        f"Attrition ({fmt_pct(attrition_pct)}): <b style='color:#ef4444'>-{attr_hc}</b> &nbsp;|&nbsp; "
        f"Net HC End of Month: <b style='color:#10b981'>{net_hc}</b>"
        f"</span></div>",
        unsafe_allow_html=True
    )

st.divider()

with st.expander("📋 Copy Month to Multiple Months"):
    cc1, cc2 = st.columns([2, 4])
    copy_from = cc1.selectbox("Copy FROM", MONTHS, index=MONTHS.index(active), key="copy_from")

    with cc2:
        st.markdown("**Copy TO** (select one or more)")
        dest_cols = st.columns(6)
        selected_targets = []
        for mi, m in enumerate(MONTHS):
            col = dest_cols[mi % 6]
            disabled = (m == copy_from)
            checked = col.checkbox(
                m,
                key=f"copy_target_{m}",
                value=False,
                disabled=disabled,
                help="Same as source" if disabled else f"Copy to {m}",
            )
            if checked and not disabled:
                selected_targets.append(m)

    st.markdown("")
    btn_col, info_col = st.columns([1, 4])
    if btn_col.button("▶ Copy", use_container_width=True, type="primary"):
        if not selected_targets:
            st.error("Please select at least one destination month.")
        else:
            for m in selected_targets:
                client()["blocks"][m] = copy.deepcopy(client()["blocks"][copy_from])
                # Copy per-month overhead override if source has one
                src_oh = client()["overhead_monthly"].get(copy_from)
                if src_oh is not None:
                    client()["overhead_monthly"][m] = copy.deepcopy(src_oh)
            targets_str = ", ".join(selected_targets)
            n_blocks = len(client()["blocks"][copy_from])
            cola_note = " COLA schedules follow block positions." if client()["cola_configs"] else ""
            st.success(f"✅ Copied **{copy_from}** ({n_blocks} blocks) → {targets_str}.{cola_note}")
            st.rerun()
    with info_col:
        if selected_targets:
            st.info(f"Will copy **{copy_from}** → {', '.join(selected_targets)}")
        else:
            st.caption("No destination months selected yet.")

st.markdown('<div class="section-title">Production Blocks</div>', unsafe_allow_html=True)
blocks = client()["blocks"][active]

if st.button("+ Add Production Block", type="secondary"):
    blocks.append({"lang":"","hc":0,"salary":0,"unit_price":0,
                   "shrink_override":None,"fx_override":None,"hours_override":None})
    st.rerun()

blocks_to_delete = []
for i, b in enumerate(blocks):
    raw_shrink = b["shrink_override"] if b.get("shrink_override") is not None else g_shrink
    shrink = max(0.0, min(0.99, raw_shrink if raw_shrink <= 1 else raw_shrink / 100))
    fx     = b["fx_override"]     if b.get("fx_override")     is not None else g_fx
    hours  = b["hours_override"]  if b.get("hours_override")  is not None else g_hours
    base_hc        = b.get("hc", 0)
    hc             = effective_hc(active, b)             # ramp-adjusted HC for display
    salary         = b.get("salary", 0)
    base_up        = b.get("unit_price", 0)
    up             = effective_up(active, i, base_up)    # COLA-adjusted
    eff            = hours * (1 - shrink)
    rev_eur        = hc * eff * up
    rev_try        = rev_eur * fx
    cost_try_total = hc * salary * g_ctc * (1 + g_bonus_pct) + hc * g_meal
    cost_e         = cost_try_total / fx if fx else 0
    margin         = rev_eur - cost_e
    margin_try     = rev_try - cost_try_total
    label  = b.get("lang") or f"Block #{i+1}"
    warn   = " ⚠️" if (hc == 0 or salary == 0) else ""
    title  = f"Block #{i+1} — {label}{warn} | HC: {hc} | Rev: {fmt_eur(rev_eur)} ({fmt_try(rev_try)}) | Margin: {fmt_eur(margin)}"

    with st.expander(title, expanded=True):
        if hc == 0:
            st.warning("⚠️ HC is 0 — this block contributes no revenue or cost.", icon="⚠️")
        if salary == 0 and hc > 0:
            st.warning("⚠️ Base salary is 0 — cost will be understated.", icon="⚠️")
        r1c1,r1c2,r1c3,r1c4,r1c5 = st.columns([2,1,2,2,1])
        new_lang = r1c1.text_input("Language / Label", value=b.get("lang",""),
                                    key=f"lang_{active}_{i}", placeholder="e.g. DE, EN, TR")
        new_hc   = r1c2.number_input("HC", value=int(b.get("hc",0)), min_value=0, step=1,
                                      key=f"hc_{active}_{i}")
        new_sal  = r1c3.number_input("Base Salary (TRY/mo)", value=float(b.get("salary",0)),
                                      min_value=0.0, step=100.0, key=f"sal_{active}_{i}")
        new_up   = r1c4.number_input("Unit Price (EUR/hr)", value=float(b.get("unit_price",0)),
                                      min_value=0.0, step=0.1, key=f"up_{active}_{i}")
        if r1c5.button("🗑 Remove", key=f"del_{active}_{i}", use_container_width=True):
            blocks_to_delete.append(i)

        r2c1,r2c2,r2c3,r2c4,r2c5,r2c6 = st.columns([2,2,2,2,2,2])
        shr_raw = r2c1.text_input(f"Shrinkage Override (global: {g_shrink*100:.0f}%)",
                                   value="" if b.get("shrink_override") is None else str(b["shrink_override"]),
                                   key=f"shr_{active}_{i}", placeholder="blank = global")
        fx_raw  = r2c2.text_input(f"FX Override (global: {g_fx})",
                                   value="" if b.get("fx_override") is None else str(b["fx_override"]),
                                   key=f"fx_{active}_{i}", placeholder="blank = global")
        hr_raw  = r2c3.text_input(f"Hours Override (global: {g_hours})",
                                   value="" if b.get("hours_override") is None else str(b["hours_override"]),
                                   key=f"hr_{active}_{i}", placeholder="blank = global")
        att_raw = r2c4.text_input(f"Attrition Override (global: {attrition_pct*100:.1f}%)",
                                   value="" if b.get("attrition_override") is None else str(b["attrition_override"]),
                                   key=f"att_{active}_{i}", placeholder="blank = global",
                                   help="Override attrition rate for this block only e.g. 0.08 for 8%")

        # ── COLA / UP increase ────────────────────────────────
        cola_key = str(i)
        cola_cfg = client()["cola_configs"].get(cola_key, {})
        with st.expander("📈 COLA / Unit Price Change", expanded=bool(cola_cfg.get("date"))):
            cc1, cc2, cc3 = st.columns([2, 2, 1])
            cola_date_val = cola_cfg.get("date", "")
            cola_up_val   = float(cola_cfg.get("new_up", up)) if cola_cfg.get("new_up") else float(b.get("unit_price", 0))
            new_cola_date = cc1.text_input("Effective date (YYYY-MM-DD)",
                                            value=cola_date_val, key=f"cola_date_{active}_{i}",
                                            placeholder="e.g. 2025-04-15",
                                            help="New UP applies from this date. Month of change is prorated.")
            new_cola_up   = cc2.number_input("New Unit Price (EUR/hr)",
                                              value=cola_up_val, step=0.1, min_value=0.0,
                                              key=f"cola_up_{active}_{i}")
            if cc3.button("Clear COLA", key=f"cola_clr_{active}_{i}", use_container_width=True):
                client()["cola_configs"].pop(cola_key, None)
                st.rerun()
            if new_cola_date.strip():
                try:
                    _dt.date.fromisoformat(new_cola_date.strip())
                    client()["cola_configs"][cola_key] = {"date": new_cola_date.strip(), "new_up": new_cola_up}
                    # Show proration preview for current month
                    eff = effective_up(active, i, b.get("unit_price", 0))
                    if eff != b.get("unit_price", 0):
                        st.caption(f"⚡ Effective UP this month: **€{eff:.4f}/hr** (prorated from €{b.get('unit_price',0):.2f} → €{new_cola_up:.2f} on {new_cola_date.strip()})")
                    else:
                        cola_dt = _dt.date.fromisoformat(new_cola_date.strip())
                        m_idx   = MONTHS.index(active) + 1
                        if cola_dt.month > m_idx:
                            st.caption(f"ℹ️ COLA not yet active this month — full new UP applies from {MONTHS[cola_dt.month-1]}")
                        else:
                            st.caption(f"✅ Full new UP €{new_cola_up:.2f}/hr active this month")
                except ValueError:
                    st.warning("Invalid date format — use YYYY-MM-DD")

        # ── Cost breakdown ────────────────────────────────────
        ctc_cost_try     = hc * salary * g_ctc * (1 + g_bonus_pct)
        meal_cost_try    = hc * g_meal
        # per-block backfill
        # Per-block attrition override or global
        raw_att          = b.get("attrition_override")
        blk_att          = raw_att if raw_att is not None else st.session_state.attrition_rate
        blk_att          = max(0.0, min(1.0, blk_att if blk_att <= 1 else blk_att / 100))
        b_hc             = hc * blk_att
        b_cost_try       = b_hc * salary * g_ctc * (1 + g_bonus_pct) + b_hc * g_meal
        b_cost_eur       = b_cost_try / fx if fx else 0
        b_hrs            = b_hc * eff * st.session_state.backfill_efficiency
        total_cost_incl  = cost_try_total + b_cost_try
        margin_incl_eur  = rev_eur - (cost_e + b_cost_eur)
        margin_incl_try  = rev_try - total_cost_incl
        # Break-even unit price for this block
        total_billable_hrs = hc * eff
        blk_breakeven    = (cost_e + b_cost_eur) / total_billable_hrs if total_billable_hrs else 0

        st.markdown("---")
        bd1, bd2, bd3, bd4, bd5, bd6 = st.columns(6)
        with bd1:
            st.markdown("**🕐 Eff. Hrs / Agent**")
            st.markdown(f"<span style='color:#8b96b0;font-size:15px;font-weight:600'>{eff:.1f} hrs</span>", unsafe_allow_html=True)
            st.caption(f"{hours}h × (1 - {shrink*100:.0f}%)")
        with bd2:
            st.markdown("**💸 Salary CTC**")
            st.markdown(f"<span style='color:#f59e0b;font-size:15px;font-weight:600'>₺{ctc_cost_try:,.0f}</span>", unsafe_allow_html=True)
            st.caption(f"₺{salary:,.0f} × {g_ctc} × (1+{g_bonus_pct*100:.0f}%)")
        with bd3:
            st.markdown("**🍽️ Meal Cards**")
            st.markdown(f"<span style='color:#f59e0b;font-size:15px;font-weight:600'>₺{meal_cost_try:,.0f}</span>", unsafe_allow_html=True)
            st.caption(f"{hc} HC × ₺{g_meal:,.0f}")
        with bd4:
            st.markdown("**🔄 Backfill Cost**")
            st.markdown(f"<span style='color:#8b5cf6;font-size:15px;font-weight:600'>₺{b_cost_try:,.0f}</span>", unsafe_allow_html=True)
            st.markdown(f"<span style='color:#8b5cf6;font-size:13px'>{fmt_eur(b_cost_eur)}</span>", unsafe_allow_html=True)
            st.caption(f"{b_hc:.2f} HC · {b_hrs:.0f} hrs @ {st.session_state.backfill_efficiency*100:.0f}% efficiency · no revenue")
        with bd5:
            st.markdown("**💰 Total Cost**")
            st.markdown(f"<span style='color:#ef4444;font-size:15px;font-weight:600'>₺{total_cost_incl:,.0f}</span>", unsafe_allow_html=True)
            st.markdown(f"<span style='color:#ef4444;font-size:13px'>{fmt_eur(cost_e + b_cost_eur)}</span>", unsafe_allow_html=True)
            st.caption("incl. backfill")
        with bd6:
            st.markdown("**📈 Revenue**")
            st.markdown(f"<span style='color:#10b981;font-size:15px;font-weight:600'>₺{rev_try:,.0f}</span>", unsafe_allow_html=True)
            st.markdown(f"<span style='color:#10b981;font-size:13px'>{fmt_eur(rev_eur)}</span>", unsafe_allow_html=True)
            st.caption(f"{hc} HC × {eff:.1f}h × €{up}/hr")
        # Break-even insight
        be_color = "#10b981" if up >= blk_breakeven else "#ef4444"
        be_label = "✅ Above break-even" if up >= blk_breakeven else "⚠️ Below break-even"
        st.markdown(
            f"<div style='background:#12192a;border:1px solid #2a3347;border-radius:5px;"
            f"padding:6px 14px;margin-top:6px;font-size:12px;color:#8b96b0'>"
            f"Break-even price: <b style='color:{be_color}'>€{blk_breakeven:.2f}/hr</b>"
            f"&nbsp;&nbsp;·&nbsp;&nbsp;Current: <b style='color:{be_color}'>€{up:.2f}/hr</b>"
            f"&nbsp;&nbsp;·&nbsp;&nbsp;<span style='color:{be_color}'>{be_label}</span>"
            f"</div>", unsafe_allow_html=True
        )

        margin_color = "#10b981" if margin_incl_eur >= 0 else "#ef4444"
        st.markdown(
            f"<div style='background:#1e2535;border:1px solid #2a3347;border-radius:6px;"
            f"padding:10px 16px;margin-top:8px;display:flex;justify-content:space-between;align-items:center'>"
            f"<span style='color:#8b96b0;font-size:12px;font-weight:600;text-transform:uppercase;letter-spacing:0.08em'>"
            f"Gross Margin <span style='color:#8b5cf6;font-weight:400'>(incl. backfill)</span></span>"
            f"<span style='color:{margin_color};font-size:20px;font-weight:700'>{fmt_eur(margin_incl_eur)}"
            f"&nbsp;&nbsp;<span style='font-size:14px'>₺{margin_incl_try:,.0f}</span>"
            f"&nbsp;&nbsp;<span style='font-size:13px'>({fmt_pct(margin_incl_eur/rev_eur) if rev_eur else '—'})</span></span>"
            f"</div>",
            unsafe_allow_html=True
        )

        blocks[i].update({
            "lang": new_lang, "hc": new_hc, "salary": new_sal, "unit_price": new_up,
            "shrink_override":    float(shr_raw) if shr_raw.strip() else None,
            "fx_override":        float(fx_raw)  if fx_raw.strip()  else None,
            "hours_override":     float(hr_raw)  if hr_raw.strip()  else None,
            "attrition_override": float(att_raw) if att_raw.strip() else None,
        })

        # ── HC Ramp Schedule ─────────────────────────────────
        ramp = blocks[i].get("hc_ramp", {})
        has_ramp = any(ramp.get(m) is not None for m in MONTHS)

        # Derive ramp direction label for expander title
        _ramp_hcs_cur = [ramp.get(m, new_hc) for m in MONTHS]
        _first = _ramp_hcs_cur[0]; _last = _ramp_hcs_cur[-1]
        _ramp_label = ""
        if has_ramp:
            if _last > _first:   _ramp_label = " 📈 ramp-up"
            elif _last < _first: _ramp_label = " 📉 ramp-down"
            elif any(v == 0 for v in _ramp_hcs_cur): _ramp_label = " ⛔ includes zero months"
            else:                _ramp_label = " ↔ non-linear"

        with st.expander(
            f"{'📈' if not has_ramp else '🔢'} HC Ramp Schedule"
            f"{'(active' + _ramp_label + ')' if has_ramp else '(optional)'}",
            expanded=has_ramp
        ):
            st.caption(
                f"Set HC per month. Leave at base ({new_hc}) if unchanged. "
                f"**Ramp-down to 0** = no revenue, no cost for that month. "
                f"The tool automatically handles partial months — "
                f"just set the HC you expect to be billing."
            )

            # Quick-fill helpers
            qf1, qf2, qf3 = st.columns(3)
            if qf1.button("⬆ Linear ramp-up to base", key=f"ramp_up_{active}_{i}",
                          help="Starts at 0 in Jan, reaches base HC by Dec"):
                new_q = {m: max(0, round(new_hc * mi / 11)) for mi, m in enumerate(MONTHS)}
                blocks[i]["hc_ramp"] = {m: v for m, v in new_q.items() if v != new_hc}
                st.rerun()
            if qf2.button("⬇ Linear ramp-down to 0", key=f"ramp_dn_{active}_{i}",
                          help="Starts at base HC in Jan, reaches 0 by Dec"):
                new_q = {m: max(0, round(new_hc * (1 - mi / 11))) for mi, m in enumerate(MONTHS)}
                blocks[i]["hc_ramp"] = {m: v for m, v in new_q.items() if v != new_hc}
                st.rerun()
            if qf3.button("🔄 Reset to flat", key=f"ramp_reset_{active}_{i}"):
                blocks[i]["hc_ramp"] = {}
                st.rerun()

            # Per-month inputs with direction arrow indicators
            rc = st.columns(12)
            new_ramp = {}
            ramp_vals = []
            for mi, m in enumerate(MONTHS):
                with rc[mi]:
                    cur = ramp.get(m)
                    v = st.number_input(
                        m, value=int(cur) if cur is not None else new_hc,
                        min_value=0, step=1,
                        key=f"ramp_{active}_{i}_{m}",
                    )
                    ramp_vals.append(v)
                    if v != new_hc:
                        new_ramp[m] = v

            # Direction + status indicators per month
            ind_cols = st.columns(12)
            prev = new_hc
            for mi, (m, v) in enumerate(zip(MONTHS, ramp_vals)):
                with ind_cols[mi]:
                    if v == 0:
                        st.markdown("<div style='text-align:center;color:#ef4444;font-size:16px'>⛔</div>",
                                    unsafe_allow_html=True)
                        st.caption("no rev")
                    elif v > prev:
                        st.markdown("<div style='text-align:center;color:#10b981;font-size:16px'>↑</div>",
                                    unsafe_allow_html=True)
                        st.caption(f"+{v-prev}")
                    elif v < prev:
                        st.markdown("<div style='text-align:center;color:#ef4444;font-size:16px'>↓</div>",
                                    unsafe_allow_html=True)
                        st.caption(f"-{prev-v}")
                    else:
                        st.markdown("<div style='text-align:center;color:#5a6480;font-size:16px'>—</div>",
                                    unsafe_allow_html=True)
                        st.caption("stable")
                    prev = v

            blocks[i]["hc_ramp"] = new_ramp if new_ramp else {}

            # Ramp preview chart
            if has_ramp or new_ramp:
                try:
                    import plotly.graph_objects as _rgo
                    ramp_hcs = ramp_vals  # already computed above
                    # Colour: green = above base, red = below base, grey = zero
                    bar_colors = []
                    for v in ramp_hcs:
                        if v == 0:           bar_colors.append("#374151")  # dark grey = no billing
                        elif v > new_hc:     bar_colors.append("#10b981")  # green = ramp-up
                        elif v < new_hc:     bar_colors.append("#ef4444")  # red = ramp-down
                        else:                bar_colors.append("#3b82f6")  # blue = stable base
                    fig_ramp = _rgo.Figure()
                    fig_ramp.add_trace(_rgo.Bar(
                        x=MONTHS, y=ramp_hcs,
                        marker_color=bar_colors,
                        text=ramp_hcs, textposition="outside", textfont=dict(size=9),
                    ))
                    fig_ramp.add_hline(y=new_hc, line_dash="dot", line_color="#5a6480",
                                       annotation_text=f"Base HC: {new_hc}",
                                       annotation_font_color="#5a6480")
                    fig_ramp.update_layout(
                        height=200, margin=dict(l=0, r=0, t=20, b=0),
                        plot_bgcolor="#0e1420", paper_bgcolor="#0e1420",
                        font=dict(color="#8b96b0", size=10),
                        xaxis=dict(showgrid=False),
                        yaxis=dict(showgrid=True, gridcolor="#1e2535"),
                        showlegend=False,
                    )
                    st.plotly_chart(fig_ramp, use_container_width=True)

                    # Rampdown impact summary
                    zero_months = [m for m, v in zip(MONTHS, ramp_hcs) if v == 0]
                    down_months = [m for m, v in zip(MONTHS, ramp_hcs) if 0 < v < new_hc]
                    if zero_months:
                        st.markdown(
                            f"<div style='background:#1e2535;border:1px solid #ef444433;"
                            f"border-radius:5px;padding:8px 14px;font-size:12px;color:#8b96b0'>"
                            f"⛔ <b style='color:#ef4444'>No billing months:</b> {', '.join(zero_months)} — "
                            f"HC = 0, revenue = €0, cost = €0 for these months.</div>",
                            unsafe_allow_html=True
                        )
                    if down_months:
                        st.markdown(
                            f"<div style='background:#1e2535;border:1px solid #f59e0b33;"
                            f"border-radius:5px;padding:8px 14px;font-size:12px;color:#8b96b0;margin-top:4px'>"
                            f"📉 <b style='color:#f59e0b'>Partial months:</b> {', '.join(down_months)} — "
                            f"revenue and cost scale proportionally to the reduced HC.</div>",
                            unsafe_allow_html=True
                        )
                except ImportError:
                    pass

if blocks_to_delete:
    for idx in sorted(blocks_to_delete, reverse=True):
        blocks.pop(idx)
    st.rerun()

# ── Overhead Roles ───────────────────────────────────────────
st.divider()
st.markdown("### 🏢 Overhead Roles")

# Determine if this month has a per-month override
has_monthly_override = client()["overhead_monthly"].get(active) is not None
mode_label = f"📌 {active} override active" if has_monthly_override else "🌐 Global config (all months)"

oh_mode_col, oh_action_col = st.columns([3,2])
oh_mode_col.caption(f"TM / QM / OM — pure cost, not billable. Ratio auto-calculates HC from production HC. {mode_label}")

with oh_action_col:
    act1, act2 = st.columns(2)
    if act1.button("📋 Copy global → all months", use_container_width=True,
                   help="Apply current global config to every month, clearing any per-month overrides"):
        client()["overhead_monthly"] = {m: None for m in MONTHS}
        st.success("Global overhead applied to all months.")
        st.rerun()
    if has_monthly_override:
        if act2.button(f"✖ Clear {active} override", use_container_width=True,
                       help=f"Remove {active} override and fall back to global"):
            client()["overhead_monthly"][active] = None
            st.rerun()
    else:
        act2.caption("No override for this month")

# Work on global or monthly config
oh_data = get_oh_cfg(active, client())
prod_hc_now = t["hc"]

oh_cols = st.columns(3)
ROLE_META = {
    "TM": {"icon":"👥", "label":"Team Manager",       "default_ratio":10, "default_sal":55000},
    "QM": {"icon":"🎯", "label":"Quality Manager",    "default_ratio":20, "default_sal":60000},
    "OM": {"icon":"⚙️", "label":"Operations Manager", "default_ratio":50, "default_sal":80000},
}

for col, (role, meta) in zip(oh_cols, ROLE_META.items()):
    cfg = oh_data.setdefault(role, {
        "ratio": meta["default_ratio"],
        "hc_override": None,
        "salary": meta["default_sal"],
    })
    with col:
        st.markdown(f"**{meta['icon']} {role} — {meta['label']}**")
        new_sal   = st.number_input(f"Base Salary TRY/mo ({role})",
                                     value=float(cfg.get("salary", meta["default_sal"])),
                                     step=1000.0, min_value=0.0, key=f"oh_sal_{active}_{role}")
        new_ratio = st.number_input(f"Span of control (agents per {role})",
                                     value=int(cfg.get("ratio", meta["default_ratio"])),
                                     step=1, min_value=1, key=f"oh_ratio_{active}_{role}",
                                     help=f"1 {role} manages this many production agents.")
        # Auto HC = exact fraction for cost model; hired HC = ceiling (you must hire whole people)
        auto_hc_exact  = prod_hc_now / new_ratio if new_ratio else 0
        auto_hc_hired  = math.ceil(auto_hc_exact) if prod_hc_now > 0 else 0
        utilization    = (auto_hc_exact / auto_hc_hired * 100) if auto_hc_hired else 0

        override_raw = st.text_input(f"HC Override (blank = auto-ceil)",
                                      value="" if cfg.get("hc_override") is None else str(int(cfg["hc_override"])),
                                      key=f"oh_hc_{active}_{role}",
                                      placeholder=f"{auto_hc_hired} (auto)",
                                      help=f"Ratio needs {auto_hc_exact:.2f} → hire {auto_hc_hired}. Override if you're sharing this role across accounts.")

        # Hired HC: override or ceiling — cost uses exact fraction, display shows hired integer
        hc_hired  = int(float(override_raw)) if override_raw.strip() else auto_hc_hired
        hc_cost   = float(override_raw) if override_raw.strip() else auto_hc_exact  # fractional for cost
        cost_try  = hc_cost * new_sal * g_ctc * (1 + g_bonus_pct) + hc_cost * g_meal
        cost_eur  = cost_try / g_fx if g_fx else 0

        # Utilization colour: green <85%, amber 85-99%, red >=100%
        if utilization >= 100:   u_color, u_label = "#ef4444", f"{utilization:.0f}% — overstretched ⚠️"
        elif utilization >= 85:  u_color, u_label = "#f59e0b", f"{utilization:.0f}% — near capacity"
        else:                    u_color, u_label = "#10b981", f"{utilization:.0f}% utilisation"

        st.markdown(
            f"<div style='background:#1e2535;border:1px solid #2a3347;border-radius:5px;"
            f"padding:8px 12px;margin-top:4px;font-size:12px;line-height:1.8'>"
            f"<div><span style='color:#8b96b0'>Hired HC: </span>"
            f"<b style='color:#e8edf5;font-size:15px'>{hc_hired}</b>"
            f"<span style='color:#5a6480'> (ratio needs {auto_hc_exact:.2f})</span></div>"
            f"<div><span style='color:#8b96b0'>Utilisation: </span>"
            f"<b style='color:{u_color}'>{u_label}</b></div>"
            f"<div><span style='color:#8b96b0'>Cost: </span>"
            f"<b style='color:#ef4444'>₺{cost_try:,.0f}</b>"
            f"<span style='color:#5a6480'> / </span>"
            f"<b style='color:#ef4444'>€{cost_eur:,.0f}</b></div>"
            f"</div>", unsafe_allow_html=True
        )
        # Detect if user changed anything vs stored global
        stored     = client()["overhead_global"].get(role, {})
        hc_new     = float(hc_hired) if override_raw.strip() else None
        changed    = (new_sal   != stored.get("salary",   meta["default_sal"])  or
                      new_ratio != stored.get("ratio",    meta["default_ratio"]) or
                      hc_new    != stored.get("hc_override"))

        if changed:
            # Write to global so all months reflect the change
            client()["overhead_global"][role] = {
                "salary": new_sal, "ratio": new_ratio, "hc_override": hc_new
            }
            # If this month had a per-month override, update that too
            if client()["overhead_monthly"].get(active) is not None:
                client()["overhead_monthly"][active][role] = {
                    "salary": new_sal, "ratio": new_ratio, "hc_override": hc_new
                }

# Overhead summary bar
oh_now = calc_overhead(active, prod_hc_now, g)
oh_total_try = oh_now["total_cost_try"]
oh_total_eur = oh_now["total_cost_eur"]
st.markdown(
    f"<div style='background:#1e2535;border:1px solid #8b5cf6;border-radius:6px;"
    f"padding:10px 18px;margin-top:8px;display:flex;justify-content:space-between;align-items:center'>"
    f"<span style='color:#8b5cf6;font-weight:600'>🏢 Total Overhead — {active}</span>"
    f"<span style='color:#e8edf5'>"
    f"TM: <b>{oh_now['TM']['hc']:.0f} HC</b> &nbsp;|&nbsp; "
    f"QM: <b>{oh_now['QM']['hc']:.0f} HC</b> &nbsp;|&nbsp; "
    f"OM: <b>{oh_now['OM']['hc']:.0f} HC</b> &nbsp;|&nbsp; "
    f"Total cost: <b style='color:#ef4444'>₺{oh_total_try:,.0f}</b> / "
    f"<b style='color:#ef4444'>€{oh_total_eur:,.0f}</b>"
    f"</span></div>", unsafe_allow_html=True
)

st.divider()
st.markdown("### 📉 P&L Summary — Full Year")

fy = {"rev":0,"rev_try":0,"cost":0,"cost_try":0,"margin":0,"margin_try":0}
month_data = {}
for m in MONTHS:
    mt = get_totals(m, g)
    month_data[m] = mt
    for k in fy: fy[k] += mt[k]

LINE_ITEMS = [
    "Revenue",
    "  Prod. Cost",
    "  Backfill Cost",
    "  Training Cost",
    "  Recruitment Cost",
    "  IT & Telephony",
    "  Facilities / Rent",
    "  CAPEX (new seats)",
    "  TM Cost",
    "  QM Cost",
    "  OM Cost",
    "Total Cost",
    "Gross Margin",
    "Margin %",
    "Break-even €/hr",
    "Avg Selling €/hr",
    "Prod HC",
    "TM HC",
    "QM HC",
    "OM HC",
    "Attrition (-)",
    "Backfill HC",
    "Net HC",
    "Billable Hrs",
    "Backfill Hrs",
    "Total Produced Hrs",
]
N = len(LINE_ITEMS)

pnl_eur = {"Line Item": LINE_ITEMS}
pnl_try = {"Line Item": LINE_ITEMS}

fy_sums = {k: 0.0 for k in ["rev","rev_try","cost","cost_try",
                              "cost_excl_backfill","backfill_cost_eur","backfill_cost_try",
                              "training_cost_eur","training_cost_try",
                              "recruitment_cost_eur","recruitment_cost_try",
                              "it_cost_eur","it_cost_try","fac_cost_eur","fac_cost_try",
                              "capex_eur","capex_try",
                              "oh_cost_eur","oh_cost_try","margin","margin_try",
                              "hrs_billable","backfill_hrs","hrs"]}
fy_oh = {"TM":{"cost_eur":0,"cost_try":0,"hc":0},
         "QM":{"cost_eur":0,"cost_try":0,"hc":0},
         "OM":{"cost_eur":0,"cost_try":0,"hc":0}}

for m in MONTHS:
    mt = month_data[m]
    for k in fy_sums:
        fy_sums[k] += mt.get(k, 0)
    for role in ("TM","QM","OM"):
        fy_oh[role]["cost_eur"] += mt["oh"][role]["cost_eur"]
        fy_oh[role]["cost_try"] += mt["oh"][role]["cost_try"]
        fy_oh[role]["hc"]       += mt["oh"][role]["hc"]

    avg_up_m   = mt["rev"] / mt["hrs_billable"] if mt["hrs_billable"] else 0
    prod_c_try = mt["cost_try"] - mt["backfill_cost_try"] - mt["oh_cost_try"]

    def row_eur(mt=mt, avg_up_m=avg_up_m):
        return [
            fmt_eur(mt["rev"]),
            fmt_eur(mt["cost_excl_backfill"]),
            fmt_eur(mt["backfill_cost_eur"]),
            fmt_eur(mt.get("training_cost_eur",0)),
            fmt_eur(mt.get("recruitment_cost_eur",0)),
            fmt_eur(mt.get("it_cost_eur",0)),
            fmt_eur(mt.get("fac_cost_eur",0)),
            fmt_eur(mt.get("capex_eur",0)),
            fmt_eur(mt["oh"]["TM"]["cost_eur"]),
            fmt_eur(mt["oh"]["QM"]["cost_eur"]),
            fmt_eur(mt["oh"]["OM"]["cost_eur"]),
            fmt_eur(mt["cost"]),
            fmt_eur(mt["margin"]),
            fmt_pct(mt["margin"]/mt["rev"]) if mt["rev"] else "—",
            f'€{mt["breakeven_up"]:.2f}',
            f'€{avg_up_m:.2f}',
            f'{mt["hc"]:.1f}',
            f'{mt["oh"]["TM"]["hc"]:.2f}',
            f'{mt["oh"]["QM"]["hc"]:.2f}',
            f'{mt["oh"]["OM"]["hc"]:.2f}',
            f'{mt["attrition_hc"]:.2f}',
            f'{mt["backfill_hc"]:.2f}',
            f'{mt["net_hc"]:.2f}',
            f'{mt["hrs_billable"]:,.0f}',
            f'{mt["backfill_hrs"]:,.0f}',
            f'{mt["hrs"]:,.0f}',
        ]

    def row_try(mt=mt, prod_c_try=prod_c_try, avg_up_m=avg_up_m):
        return [
            fmt_try(mt["rev_try"]),
            fmt_try(prod_c_try),
            fmt_try(mt["backfill_cost_try"]),
            fmt_try(mt.get("training_cost_try",0)),
            fmt_try(mt.get("recruitment_cost_try",0)),
            fmt_try(mt.get("it_cost_try",0)),
            fmt_try(mt.get("fac_cost_try",0)),
            fmt_try(mt.get("capex_try",0)),
            fmt_try(mt["oh"]["TM"]["cost_try"]),
            fmt_try(mt["oh"]["QM"]["cost_try"]),
            fmt_try(mt["oh"]["OM"]["cost_try"]),
            fmt_try(mt["cost_try"]),
            fmt_try(mt["margin_try"]),
            fmt_pct(mt["margin_try"]/mt["rev_try"]) if mt["rev_try"] else "—",
            f'€{mt["breakeven_up"]:.2f}',
            f'€{avg_up_m:.2f}',
            f'{mt["hc"]:.1f}',
            f'{mt["oh"]["TM"]["hc"]:.2f}',
            f'{mt["oh"]["QM"]["hc"]:.2f}',
            f'{mt["oh"]["OM"]["hc"]:.2f}',
            f'{mt["attrition_hc"]:.2f}',
            f'{mt["backfill_hc"]:.2f}',
            f'{mt["net_hc"]:.2f}',
            f'{mt["hrs_billable"]:,.0f}',
            f'{mt["backfill_hrs"]:,.0f}',
            f'{mt["hrs"]:,.0f}',
        ]

    r_e = row_eur(); r_t = row_try()
    assert len(r_e) == N, f"EUR row {m}: {len(r_e)} != {N}"
    assert len(r_t) == N, f"TRY row {m}: {len(r_t)} != {N}"
    pnl_eur[m] = r_e
    pnl_try[m] = r_t

# Full Year totals
fy_be     = fy_sums["cost"]     / fy_sums["hrs_billable"] if fy_sums["hrs_billable"] else 0
fy_avg_up = fy_sums["rev"]      / fy_sums["hrs_billable"] if fy_sums["hrs_billable"] else 0
fy_prod_c_try = fy_sums["cost_try"] - fy_sums["backfill_cost_try"] - fy_sums["oh_cost_try"]

def fy_row_eur():
    return [
        fmt_eur(fy_sums["rev"]),
        fmt_eur(fy_sums["cost_excl_backfill"]),
        fmt_eur(fy_sums["backfill_cost_eur"]),
        fmt_eur(fy_sums["training_cost_eur"]),
        fmt_eur(fy_sums["recruitment_cost_eur"]),
        fmt_eur(fy_sums["it_cost_eur"]),
        fmt_eur(fy_sums["fac_cost_eur"]),
        fmt_eur(fy_sums["capex_eur"]),
        fmt_eur(fy_oh["TM"]["cost_eur"]),
        fmt_eur(fy_oh["QM"]["cost_eur"]),
        fmt_eur(fy_oh["OM"]["cost_eur"]),
        fmt_eur(fy_sums["cost"]),
        fmt_eur(fy_sums["margin"]),
        fmt_pct(fy_sums["margin"]/fy_sums["rev"]) if fy_sums["rev"] else "—",
        f'€{fy_be:.2f}', f'€{fy_avg_up:.2f}',
        "","","","","","","",
        f'{fy_sums["hrs_billable"]:,.0f}',
        f'{fy_sums["backfill_hrs"]:,.0f}',
        f'{fy_sums["hrs"]:,.0f}',
    ]

def fy_row_try():
    return [
        fmt_try(fy_sums["rev_try"]),
        fmt_try(fy_prod_c_try),
        fmt_try(fy_sums["backfill_cost_try"]),
        fmt_try(fy_sums["training_cost_try"]),
        fmt_try(fy_sums["recruitment_cost_try"]),
        fmt_try(fy_sums["it_cost_try"]),
        fmt_try(fy_sums["fac_cost_try"]),
        fmt_try(fy_sums["capex_try"]),
        fmt_try(fy_oh["TM"]["cost_try"]),
        fmt_try(fy_oh["QM"]["cost_try"]),
        fmt_try(fy_oh["OM"]["cost_try"]),
        fmt_try(fy_sums["cost_try"]),
        fmt_try(fy_sums["margin_try"]),
        fmt_pct(fy_sums["margin_try"]/fy_sums["rev_try"]) if fy_sums["rev_try"] else "—",
        f'€{fy_be:.2f}', f'€{fy_avg_up:.2f}',
        "","","","","","","",
        f'{fy_sums["hrs_billable"]:,.0f}',
        f'{fy_sums["backfill_hrs"]:,.0f}',
        f'{fy_sums["hrs"]:,.0f}',
    ]

r_fy_e = fy_row_eur(); r_fy_t = fy_row_try()
assert len(r_fy_e) == N, f"FY EUR: {len(r_fy_e)} != {N}"
assert len(r_fy_t) == N, f"FY TRY: {len(r_fy_t)} != {N}"
pnl_eur["Full Year"] = r_fy_e
pnl_try["Full Year"] = r_fy_t

tab_eur, tab_try = st.tabs(["💶 EUR View", "₺ TRY View"])
with tab_eur:
    st.dataframe(pd.DataFrame(pnl_eur).set_index("Line Item"), use_container_width=True)
with tab_try:
    st.dataframe(pd.DataFrame(pnl_try).set_index("Line Item"), use_container_width=True)

# ── Actual vs Budget ─────────────────────────────────────────"Enter monthly actuals to track variance against your budget. All figures in EUR.")

with st.expander("✏️ Enter / Edit Actuals", expanded=False):
    for row_months in [MONTHS[:6], MONTHS[6:]]:
        cols = st.columns(6)
        for col, m in zip(cols, row_months):
            act = client()["actuals"].get(m, {})
            with col:
                st.markdown(f"**{m}**")
                a_rev  = st.number_input(f"Rev €",  value=float(act.get("rev",0)),  step=100.0, min_value=0.0, key=f"act_rev_{m}")
                a_cost = st.number_input(f"Cost €", value=float(act.get("cost",0)), step=100.0, min_value=0.0, key=f"act_cost_{m}")
                a_hc   = st.number_input(f"HC",     value=int(act.get("hc",0)),     step=1,     min_value=0,   key=f"act_hc_{m}")
                client()["actuals"][m] = {"rev": a_rev, "cost": a_cost, "hc": a_hc, "margin": a_rev - a_cost}

has_any_actual = any(bool(client()["actuals"].get(m,{}).get("rev") or
                          client()["actuals"].get(m,{}).get("cost")) for m in MONTHS)

if has_any_actual:
    months_with_actuals = [m for m in MONTHS if client()["actuals"].get(m,{}).get("rev") or
                                                client()["actuals"].get(m,{}).get("cost")]
    bgt_ytd = sum(month_data[m]["rev"]    for m in months_with_actuals)
    act_ytd = sum(client()["actuals"].get(m,{}).get("rev",0) for m in months_with_actuals)
    bgm_ytd = sum(month_data[m]["margin"] for m in months_with_actuals)
    agm_ytd = sum(client()["actuals"].get(m,{}).get("margin",0) for m in months_with_actuals)
    rev_ytd_var = act_ytd - bgt_ytd
    gm_ytd_var  = agm_ytd - bgm_ytd
    rv_color = "#10b981" if rev_ytd_var >= 0 else "#ef4444"
    gv_color = "#10b981" if gm_ytd_var  >= 0 else "#ef4444"
    n_months = len(months_with_actuals)
    st.markdown(
        f"<div style='background:#1e2535;border:1px solid #2a3347;border-radius:6px;"
        f"padding:10px 20px;margin-bottom:12px;display:flex;gap:40px;align-items:center'>"
        f"<span style='color:#8b96b0;font-size:12px'>YTD ({n_months} months reported)</span>"
        f"<span style='color:#e8edf5'>Revenue: <b>{fmt_eur(act_ytd)}</b> vs <b>{fmt_eur(bgt_ytd)}</b> bgt "
        f"<b style='color:{rv_color}'>({'+' if rev_ytd_var>=0 else ''}{fmt_eur(rev_ytd_var)})</b></span>"
        f"<span style='color:#e8edf5'>Gross Margin: <b>{fmt_eur(agm_ytd)}</b> vs <b>{fmt_eur(bgm_ytd)}</b> bgt "
        f"<b style='color:{gv_color}'>({'+' if gm_ytd_var>=0 else ''}{fmt_eur(gm_ytd_var)})</b></span>"
        f"</div>", unsafe_allow_html=True
    )

    avb_rows = []
    for m in MONTHS:
        mt  = month_data[m]
        act = client()["actuals"].get(m, {})
        has_act = bool(act.get("rev") or act.get("cost"))
        b_rev=mt["rev"]; a_rev=act.get("rev",0)
        b_gm=mt["margin"]; a_gm=act.get("margin", a_rev - act.get("cost",0))
        b_cost=mt["cost"]; a_cost=act.get("cost",0)
        rv = a_rev-b_rev if has_act else None
        gv = a_gm-b_gm   if has_act else None
        cv = a_cost-b_cost if has_act else None
        def fv(v, pos_good=True):
            if v is None: return "—"
            s = f"{'+' if v>=0 else ''}€{abs(v):,.0f}" if v>=0 else f"-€{abs(v):,.0f}"
            return s
        def fp(v, base):
            if v is None or not base: return "—"
            pct = v/base*100
            return f"{'+' if pct>=0 else ''}{pct:.1f}%"
        avb_rows.append({
            "Month":m,
            "Bgt Rev":fmt_eur(b_rev), "Act Rev":fmt_eur(a_rev) if has_act else "—",
            "Rev Var":fv(rv), "Rev Var%":fp(rv,b_rev),
            "Bgt Cost":fmt_eur(b_cost), "Act Cost":fmt_eur(a_cost) if has_act else "—",
            "Cost Var":fv(cv, pos_good=False),
            "Bgt GM":fmt_eur(b_gm), "Act GM":fmt_eur(a_gm) if has_act else "—",
            "GM Var":fv(gv), "GM Var%":fp(gv,b_gm),
            "Bgt HC":f"{int(mt['hc'])}", "Act HC":f"{int(act.get('hc',0))}" if has_act else "—",
        })
    df_avb = pd.DataFrame(avb_rows).set_index("Month")
    st.dataframe(df_avb, use_container_width=True)

    try:
        import plotly.graph_objects as _go
        months_act = [m for m in MONTHS if client()["actuals"].get(m,{}).get("rev") or
                                           client()["actuals"].get(m,{}).get("cost")]
        rev_vars_ch = [client()["actuals"].get(m,{}).get("rev",0) - month_data[m]["rev"] for m in months_act]
        gm_vars_ch  = [client()["actuals"].get(m,{}).get("margin",0) - month_data[m]["margin"] for m in months_act]
        fig_avb = _go.Figure()
        fig_avb.add_trace(_go.Bar(name="Revenue Variance", x=months_act, y=rev_vars_ch,
            marker_color=["#10b981" if v>=0 else "#ef4444" for v in rev_vars_ch],
            text=[f"{'+' if v>=0 else ''}€{v:,.0f}" for v in rev_vars_ch], textposition="outside", textfont=dict(size=10)))
        fig_avb.add_trace(_go.Bar(name="GM Variance", x=months_act, y=gm_vars_ch,
            marker_color=["#3b82f6" if v>=0 else "#f59e0b" for v in gm_vars_ch],
            text=[f"{'+' if v>=0 else ''}€{v:,.0f}" for v in gm_vars_ch], textposition="outside", textfont=dict(size=10)))
        fig_avb.add_hline(y=0, line_color="#2a3347", line_width=1.5)
        fig_avb.update_layout(
            title=dict(text="Actual vs Budget Variance (EUR)", font=dict(color="#e8edf5", size=13)),
            barmode="group", bargap=0.2, plot_bgcolor="#0e1420", paper_bgcolor="#0e1420",
            font=dict(color="#8b96b0"), legend=dict(orientation="h", y=1.08, bgcolor="rgba(0,0,0,0)"),
            margin=dict(l=10, r=10, t=50, b=10), height=340,
            xaxis=dict(showgrid=False),
            yaxis=dict(showgrid=True, gridcolor="#1e2535", tickprefix="€", zeroline=False),
            hoverlabel=dict(bgcolor="#1e2535", bordercolor="#2a3347", font=dict(color="#e8edf5")),
        )
        st.plotly_chart(fig_avb, use_container_width=True)
    except ImportError:
        pass
else:
    st.info("No actuals entered yet. Use '✏️ Enter / Edit Actuals' above to start tracking.", icon="ℹ️")

# ── FX Scenario Projection ───────────────────────────────────
st.divider()
st.markdown("### 💱 FX Rate Projection — EUR/TRY")
st.caption("Model how TRY depreciation affects your full-year cost base. Uses live rate as anchor.")

_live_fx, _fx_ok = fetch_live_fx()
_fx_anchor = g["fx"]  # use the rate currently set in sidebar (may be manual)

fxp_c1, fxp_c2, fxp_c3, fxp_c4 = st.columns(4)
fxp_bear = fxp_c1.number_input("🐻 Bear rate (yr-end)",
    value=round(_fx_anchor * 1.20, 1), step=0.5,
    help="Worst case: TRY depreciates 20%+ vs EUR by Dec.")
fxp_base = fxp_c2.number_input("📊 Base rate (yr-end)",
    value=round(_fx_anchor * 1.10, 1), step=0.5,
    help="Most likely: ~10% annual TRY depreciation.")
fxp_bull = fxp_c3.number_input("🐂 Bull rate (yr-end)",
    value=round(_fx_anchor * 1.02, 1), step=0.5,
    help="Optimistic: TRY nearly stable vs EUR.")
fxp_now  = fxp_c4.number_input("📍 Current / anchor rate",
    value=float(_fx_anchor), step=0.5,
    help="Starting point. Defaults to sidebar FX rate.")

# Build month-by-month linear interpolation for each scenario
def _interp_fx(start, end):
    """Linear interpolation from start (Jan) to end (Dec) across 12 months."""
    return [round(start + (end - start) * i / 11, 2) for i in range(12)]

proj_bear = _interp_fx(fxp_now, fxp_bear)
proj_base = _interp_fx(fxp_now, fxp_base)
proj_bull = _interp_fx(fxp_now, fxp_bull)

# Show projected rates table
proj_df = pd.DataFrame({
    "Month":      MONTHS,
    "🐻 Bear":    proj_bear,
    "📊 Base":    proj_base,
    "🐂 Bull":    proj_bull,
})

# Compute cost impact: recalc total cost at each scenario FX vs current budget FX
# Cost in EUR = cost_TRY / FX  → higher FX = lower EUR cost (TRY cheaper)
impact_rows = []
for mi, m in enumerate(MONTHS):
    mt = month_data[m]
    cost_try = mt["cost_try"]
    bgt_fx   = _fx_anchor
    for scen, proj in [("Bear", proj_bear), ("Base", proj_base), ("Bull", proj_bull)]:
        scen_fx  = proj[mi]
        scen_cost_eur = cost_try / scen_fx if scen_fx else 0
        delta_eur     = scen_cost_eur - mt["cost"]   # vs budget cost
        impact_rows.append({"Month": m, "Scenario": scen,
                             "FX": scen_fx, "Cost EUR": scen_cost_eur,
                             "Δ vs Budget": delta_eur})

try:
    import plotly.graph_objects as _fxgo
    fig_fx = _fxgo.Figure()

    _scenarios = [
        ("🐻 Bear", proj_bear, "#ef4444", "rgba(239,68,68,0.08)"),
        ("📊 Base", proj_base, "#3b82f6", "rgba(59,130,246,0.12)"),
        ("🐂 Bull", proj_bull, "#10b981", "rgba(16,185,129,0.08)"),
    ]
    for label, proj, color, fill in _scenarios:
        fig_fx.add_trace(_fxgo.Scatter(
            name=label, x=MONTHS, y=proj,
            mode="lines+markers",
            line=dict(color=color, width=2.5),
            marker=dict(size=6, color=color),
            fill="tozeroy", fillcolor=fill,
            hovertemplate=f"{label}<br>%{{x}}: ₺%{{y:,.2f}}<extra></extra>",
        ))

    # Horizontal line for current rate
    fig_fx.add_hline(y=fxp_now, line_dash="dot", line_color="#5a6480",
                     annotation_text=f"Current: ₺{fxp_now:,.2f}",
                     annotation_font_color="#5a6480", annotation_position="right")

    fig_fx.update_layout(
        plot_bgcolor="#0e1420", paper_bgcolor="#0e1420",
        font=dict(color="#8b96b0", family="Inter, sans-serif"),
        legend=dict(orientation="h", y=1.06, bgcolor="rgba(0,0,0,0)", font=dict(size=11)),
        margin=dict(l=10, r=80, t=30, b=10), height=300,
        xaxis=dict(showgrid=False, tickfont=dict(color="#8b96b0")),
        yaxis=dict(showgrid=True, gridcolor="#1e2535",
                   tickprefix="₺", tickfont=dict(color="#8b96b0")),
        hoverlabel=dict(bgcolor="#1e2535", bordercolor="#2a3347", font=dict(color="#e8edf5")),
    )
    st.plotly_chart(fig_fx, use_container_width=True)

    # Cost impact chart
    st.markdown("**💸 Extra margin (or loss) from FX movement — vs your budget rate**")
    st.caption(
        "Shows how much your EUR cost changes each month compared to what you budgeted, "
        "purely due to TRY/EUR movement. "
        "📉 **Negative bar = your costs are CHEAPER in EUR → extra margin in your pocket.** "
        "📈 Positive bar = TRY strengthened → costs are higher than planned. "
        "Since you pay salaries in TRY but bill in EUR, a weaker TRY always helps your margin."
    )
    fig_imp = _fxgo.Figure()
    for scen, color in [("Bear","#ef4444"),("Base","#3b82f6"),("Bull","#10b981")]:
        rows = [r for r in impact_rows if r["Scenario"] == scen]
        deltas = [r["Δ vs Budget"] for r in rows]
        fig_imp.add_trace(_fxgo.Bar(
            name=f"{'🐻 Bear — TRY weakens most' if scen=='Bear' else '📊 Base — moderate depreciation' if scen=='Base' else '🐂 Bull — TRY stays strong'}",
            x=MONTHS, y=deltas,
            marker_color=color,
            opacity=0.85,
            hovertemplate=(
                "<b>%{x}</b><br>"
                "FX delta vs budget: <b>%{y:+,.0f} EUR</b><br>"
                "<i>Negative = cheaper costs = extra margin</i>"
                "<extra></extra>"
            ),
        ))
    fig_imp.add_hline(y=0, line_color="#5a6480", line_width=1,
                      annotation_text="Budget FX baseline (€0 impact)",
                      annotation_font_color="#5a6480",
                      annotation_position="top left")
    fig_imp.update_layout(
        barmode="group", bargap=0.15,
        plot_bgcolor="#0e1420", paper_bgcolor="#0e1420",
        font=dict(color="#8b96b0"),
        legend=dict(
            orientation="h",
            yanchor="bottom", y=1.12,   # push well above bars
            xanchor="center", x=0.5,
            bgcolor="rgba(14,20,32,0.8)",
            bordercolor="#2a3347", borderwidth=1,
            font=dict(size=11),
        ),
        margin=dict(l=10, r=10, t=70, b=10), height=310,  # extra top margin for legend
        xaxis=dict(showgrid=False),
        yaxis=dict(showgrid=True, gridcolor="#1e2535",
                   tickprefix="€", zeroline=False,
                   title=dict(text="Cost delta vs budget", font=dict(size=10, color="#5a6480"))),
        hoverlabel=dict(bgcolor="#1e2535", bordercolor="#2a3347", font=dict(color="#e8edf5")),
    )
    st.plotly_chart(fig_imp, use_container_width=True)

    # Summary table: full-year cost impact per scenario
    st.markdown("**📊 Full-year cost summary by FX scenario**")
    fy_impact = {}
    for scen in ["Bear","Base","Bull"]:
        rows = [r for r in impact_rows if r["Scenario"] == scen]
        fy_cost  = sum(r["Cost EUR"] for r in rows)
        fy_delta = sum(r["Δ vs Budget"] for r in rows)
        fy_rev   = sum(month_data[m]["rev"] for m in MONTHS)
        fy_impact[scen] = {"FY Cost EUR": fy_cost, "Δ vs Budget": fy_delta,
                           "FY Margin": fy_rev - fy_cost,
                           "Margin %": (fy_rev - fy_cost)/fy_rev*100 if fy_rev else 0}

    imp_cols = st.columns(3)
    icons = {"Bear":"🐻","Base":"📊","Bull":"🐂"}
    colors_scen = {"Bear":"#ef4444","Base":"#3b82f6","Bull":"#10b981"}
    for col, scen in zip(imp_cols, ["Bear","Base","Bull"]):
        d = fy_impact[scen]
        delta_color = "#10b981" if d["Δ vs Budget"] <= 0 else "#ef4444"
        col.markdown(
            f"<div style='background:#1e2535;border:1px solid {colors_scen[scen]}33;"
            f"border-radius:8px;padding:14px 16px;text-align:center'>"
            f"<div style='color:{colors_scen[scen]};font-weight:700;font-size:14px;margin-bottom:8px'>"
            f"{icons[scen]} {scen} Case  ·  yr-end ₺{proj_bear[-1] if scen=='Bear' else proj_base[-1] if scen=='Base' else proj_bull[-1]:,.1f}</div>"
            f"<div style='color:#e8edf5;font-size:18px;font-weight:700'>€{d['FY Cost EUR']:,.0f}</div>"
            f"<div style='color:#8b96b0;font-size:12px'>FY Total Cost</div>"
            f"<div style='color:{delta_color};font-size:13px;margin-top:6px'>"
            f"{'↓' if d['Δ vs Budget']<=0 else '↑'} €{abs(d['Δ vs Budget']):,.0f} vs budget FX</div>"
            f"<div style='color:#8b96b0;font-size:12px;margin-top:4px'>"
            f"Margin: <b style='color:#e8edf5'>{d['Margin %']:.1f}%</b></div>"
            f"</div>", unsafe_allow_html=True
        )

except ImportError:
    st.info("Install plotly to see FX projection charts.")

# ── Charts ───────────────────────────────────────────────────
st.divider()
st.markdown("### 📊 Performance Charts")

try:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots

    chart_months = MONTHS
    # Guard: month_data may be empty if no blocks configured
    if not month_data or all(month_data[m]["hc"] == 0 for m in MONTHS):
        st.info("Add production blocks to see charts.", icon="📊")
        raise ImportError("no data")  # skip chart rendering gracefully
    revs   = [month_data[m]["rev"]    for m in chart_months]
    costs  = [month_data[m]["cost"]   for m in chart_months]
    gms    = [month_data[m]["margin"] for m in chart_months]
    margins= [month_data[m]["margin"] / month_data[m]["rev"] * 100
              if month_data[m]["rev"] else 0 for m in chart_months]
    be_ups = [month_data[m]["breakeven_up"] for m in chart_months]
    avg_ups= [month_data[m]["rev"] / month_data[m]["hrs_billable"]
              if month_data[m]["hrs_billable"] else 0 for m in chart_months]

    # ── Chart 1: Revenue / Cost / GM bars + Margin% line ─────
    fig = make_subplots(specs=[[{"secondary_y": True}]])

    bar_w = 0.25
    fig.add_trace(go.Bar(
        name="Revenue (EUR)", x=chart_months, y=revs,
        marker_color="#3b82f6", opacity=0.85,
        text=[f"€{v/1000:.0f}k" for v in revs], textposition="outside",
        textfont=dict(size=10, color="#8b96b0"),
    ), secondary_y=False)
    fig.add_trace(go.Bar(
        name="Total Cost (EUR)", x=chart_months, y=costs,
        marker_color="#ef4444", opacity=0.85,
        text=[f"€{v/1000:.0f}k" for v in costs], textposition="outside",
        textfont=dict(size=10, color="#8b96b0"),
    ), secondary_y=False)
    fig.add_trace(go.Bar(
        name="Gross Margin (EUR)", x=chart_months, y=gms,
        marker_color="#10b981", opacity=0.85,
        text=[f"€{v/1000:.0f}k" for v in gms], textposition="outside",
        textfont=dict(size=10, color="#8b96b0"),
    ), secondary_y=False)
    fig.add_trace(go.Scatter(
        name="Margin %", x=chart_months, y=margins,
        mode="lines+markers+text",
        line=dict(color="#f59e0b", width=2.5, dash="dot"),
        marker=dict(size=7, color="#f59e0b",
                    line=dict(color="#1e2535", width=2)),
        text=[f"{v:.1f}%" for v in margins],
        textposition="top center",
        textfont=dict(size=10, color="#f59e0b"),
        yaxis="y2",
    ), secondary_y=True)

    fig.update_layout(
        barmode="group", bargap=0.18, bargroupgap=0.05,
        plot_bgcolor="#0e1420", paper_bgcolor="#0e1420",
        font=dict(color="#8b96b0", family="Inter, sans-serif"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02,
                    xanchor="right", x=1,
                    bgcolor="rgba(0,0,0,0)", font=dict(size=11)),
        margin=dict(l=10, r=10, t=40, b=10),
        height=420,
        xaxis=dict(showgrid=False, tickfont=dict(color="#8b96b0")),
        yaxis=dict(showgrid=True, gridcolor="#1e2535",
                   tickprefix="€", tickfont=dict(color="#8b96b0"), title=""),
        yaxis2=dict(showgrid=False, ticksuffix="%",
                    tickfont=dict(color="#f59e0b"),
                    range=[0, max(margins) * 1.4 if any(m>0 for m in margins) else 100],
                    title=""),
        hoverlabel=dict(bgcolor="#1e2535", bordercolor="#2a3347",
                        font=dict(color="#e8edf5")),
    )
    fig.update_traces(hovertemplate="%{x}<br>%{y:,.0f}<extra>%{fullData.name}</extra>")
    st.plotly_chart(fig, use_container_width=True)

    # ── Chart 2: Break-even vs Avg Selling Price ──────────────
    fig2 = go.Figure()
    fig2.add_trace(go.Scatter(
        name="Avg Selling Price (€/hr)", x=chart_months, y=avg_ups,
        mode="lines+markers",
        line=dict(color="#3b82f6", width=2.5),
        marker=dict(size=8, color="#3b82f6", line=dict(color="#1e2535", width=2)),
        fill="tozeroy", fillcolor="rgba(59,130,246,0.08)",
    ))
    fig2.add_trace(go.Scatter(
        name="Break-even Price (€/hr)", x=chart_months, y=be_ups,
        mode="lines+markers",
        line=dict(color="#ef4444", width=2, dash="dash"),
        marker=dict(size=7, color="#ef4444", line=dict(color="#1e2535", width=2)),
        fill="tozeroy", fillcolor="rgba(239,68,68,0.05)",
    ))
    # Shade gap between the two lines
    fig2.add_trace(go.Scatter(
        name="Margin buffer (€/hr)",
        x=chart_months + chart_months[::-1],
        y=avg_ups + be_ups[::-1],
        fill="toself",
        fillcolor="rgba(16,185,129,0.08)",
        line=dict(color="rgba(0,0,0,0)"),
        showlegend=False, hoverinfo="skip",
    ))
    fig2.update_layout(
        plot_bgcolor="#0e1420", paper_bgcolor="#0e1420",
        font=dict(color="#8b96b0", family="Inter, sans-serif"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02,
                    xanchor="right", x=1, bgcolor="rgba(0,0,0,0)",
                    font=dict(size=11)),
        margin=dict(l=10, r=10, t=40, b=10),
        height=300,
        xaxis=dict(showgrid=False, tickfont=dict(color="#8b96b0")),
        yaxis=dict(showgrid=True, gridcolor="#1e2535",
                   tickprefix="€", ticksuffix="/hr",
                   tickfont=dict(color="#8b96b0"), title=""),
        hoverlabel=dict(bgcolor="#1e2535", bordercolor="#2a3347",
                        font=dict(color="#e8edf5")),
    )
    st.plotly_chart(fig2, use_container_width=True)

    # ── HC Ramp overview chart ───────────────────────────────
    any_ramp = any(
        any(b.get("hc_ramp") for b in client()["blocks"].get(m, []))
        for m in MONTHS
    )
    if any_ramp:
        st.markdown("#### 👥 HC Ramp Overview")
        all_blocks_labels = []
        for m_blks in client()["blocks"].values():
            for b in m_blks:
                lbl = b.get("lang") or "Block"
                if lbl not in all_blocks_labels:
                    all_blocks_labels.append(lbl)

        fig_hc = go.Figure()
        # Total HC per month (all blocks combined, ramp-adjusted)
        total_hcs = [sum(effective_hc(m, b) for b in client()["blocks"].get(m, [])) for m in MONTHS]
        fig_hc.add_trace(go.Scatter(
            name="Total HC", x=MONTHS, y=total_hcs,
            mode="lines+markers+text",
            line=dict(color="#3b82f6", width=2.5),
            marker=dict(size=7, color="#3b82f6"),
            text=total_hcs, textposition="top center", textfont=dict(size=9),
            fill="tozeroy", fillcolor="rgba(59,130,246,0.06)",
        ))
        fig_hc.update_layout(
            plot_bgcolor="#0e1420", paper_bgcolor="#0e1420",
            font=dict(color="#8b96b0"),
            margin=dict(l=10, r=10, t=20, b=10), height=260,
            xaxis=dict(showgrid=False),
            yaxis=dict(showgrid=True, gridcolor="#1e2535", title=""),
            showlegend=False,
            hoverlabel=dict(bgcolor="#1e2535", bordercolor="#2a3347"),
        )
        st.plotly_chart(fig_hc, use_container_width=True)

except ImportError:
    st.info("Install plotly for charts: `pip install plotly`")

st.divider()

# ── Formula Reference ─────────────────────────────────────────
with st.expander("📐 Formula Reference — How Calculations Work", expanded=False):

    st.markdown("""
<style>
.formula-section { margin-bottom: 24px; }
.formula-title {
    font-size: 12px; font-weight: 700; letter-spacing: 0.08em;
    text-transform: uppercase; color: #5a6480;
    border-bottom: 1px solid #2a3347;
    padding-bottom: 4px; margin-bottom: 10px;
}
.formula-row {
    display: grid; grid-template-columns: 220px 1fr;
    gap: 8px; align-items: start;
    padding: 6px 0; border-bottom: 1px solid #1e2535;
    font-size: 13px;
}
.formula-label { color: #8b96b0; font-weight: 600; }
.formula-expr  { color: #e8edf5; font-family: monospace; font-size: 12px; }
.formula-note  { color: #5a6480; font-size: 11px; margin-top: 2px; }
</style>
""", unsafe_allow_html=True)

    def section(title):
        st.markdown(f"<div class='formula-title'>{title}</div>", unsafe_allow_html=True)

    def row(label, expr, note=""):
        note_html = f"<div class='formula-note'>{note}</div>" if note else ""
        st.markdown(
            f"<div class='formula-row'>"
            f"<div class='formula-label'>{label}</div>"
            f"<div><div class='formula-expr'>{expr}</div>{note_html}</div>"
            f"</div>", unsafe_allow_html=True)

    col_a, col_b = st.columns(2)

    with col_a:
        section("⏱ Hours & Productivity")
        row("Gross hours / agent",    "hours_per_month",
            "Global setting. Default 160 hrs/mo.")
        row("Shrinkage",              "shrink %  (global or per-block override)",
            "Accounts for breaks, sick leave, training, etc.")
        row("Effective hours / agent","gross_hours × (1 − shrink)",
            "Billable productive hours per agent per month.")
        row("Total billable hours",   "Σ (HC × effective_hrs)  across all blocks",
            "Sum across all production blocks for the month.")
        row("COLA-adjusted UP",       "base_UP  if month < COLA date",
            "Prorated in transition month: (days_old×old_UP + days_new×new_UP) ÷ days_in_month")
        row("HC ramp",                "hc_ramp[month]  if set,  else block base HC",
            "Per-block monthly override. Blank = use base HC.")

        st.markdown("<br>", unsafe_allow_html=True)
        section("💰 Revenue")
        row("Block revenue (EUR)",    "HC × effective_hrs × COLA-adjusted UP",
            "UP = Unit Price in EUR/hr.")
        row("Block revenue (TRY)",    "revenue_EUR × FX rate",
            "FX = global rate or per-block override.")
        row("Total revenue",          "Σ block revenues across all blocks")

        st.markdown("<br>", unsafe_allow_html=True)
        section("💸 Production Cost")
        row("Gross salary cost (TRY)","HC × base_salary × CTC_multiplier × (1 + bonus %)",
            "CTC = Cost to Company multiplier. Covers employer taxes, insurance, etc.")
        row("Meal card cost (TRY)",   "HC × meal_card_TRY / month",
            "Fixed per-head monthly benefit.")
        row("Total prod cost (TRY)",  "salary_cost + meal_cost")
        row("Total prod cost (EUR)",  "total_cost_TRY ÷ FX rate")

        st.markdown("<br>", unsafe_allow_html=True)
        section("🔄 Attrition & Backfill")
        row("Attrition HC",           "Σ (block_HC × block_attrition_rate)",
            "Per-block rate if set, otherwise global attrition %.")
        row("Backfill HC",            "= attrition HC  (1-for-1 replacement)")
        row("Net HC (EOM)",           "total HC − attrition HC")
        row("Backfill salary cost",   "backfill_HC × avg_salary × CTC × (1 + bonus%) + backfill_HC × meal",
            "Backfill agents earn full salary from day 1.")
        row("Backfill hours",         "backfill_HC × avg_eff_hrs × training_efficiency %",
            "Partial productivity during onboarding. Hours produced but NOT billed.")

    with col_b:
        section("🏢 Overhead Roles (TM / QM / OM)")
        row("Auto HC (exact)",        "prod_HC ÷ span_of_control",
            "Fractional — used for cost calculation only.")
        row("Hired HC (ceiling)",     "⌈ prod_HC ÷ span_of_control ⌉",
            "You hire whole people. math.ceil() applied.")
        row("Overhead cost (TRY)",    "hired_HC × role_salary × CTC × (1 + bonus%) + hired_HC × meal",
            "HC override replaces auto-ceil if set.")
        row("Overhead cost (EUR)",    "overhead_cost_TRY ÷ FX rate")
        row("Utilisation %",          "(exact HC ÷ hired HC) × 100",
            "🟢 < 85%  🟡 85-99%  🔴 ≥ 100% (overstretched)")

        st.markdown("<br>", unsafe_allow_html=True)
        section("📦 OPEX & CAPEX")
        row("Training cost (TRY)",       "backfill_HC × training_cost_per_hire",
            "One-time. Triggered by attrition backfill only.")
        row("Recruitment fee (TRY)",     "(backfill_HC + HC_increase) × recruitment_fee",
            "One-time per new person. Triggered by BOTH backfill AND ramp-up delta.")
        row("IT & Telephony (TRY)",      "active_HC × it_cost_per_seat",
            "Recurring monthly. Scales up/down with every HC change.")
        row("Facilities / Rent (TRY)",   "active_HC × facilities_per_seat",
            "Recurring monthly. Desk cost: rent + electricity + cleaning per seat.")
        row("CAPEX per new seat (TRY)",  "HC_increase × (capex_pc + capex_headset + capex_software)",
            "One-time. Only fires when HC grows vs prior month. No CAPEX on backfill or stable/down months.")
        row("HC increase (delta)",       "max(0, this_month_HC − prior_month_HC)",
            "Negative delta (ramp-down) = zero CAPEX. CAPEX only on net new seats.")
        row("All OPEX/CAPEX → EUR",      "cost_TRY ÷ weighted_avg_FX",
            "Same FX conversion as all other TRY costs.")

        st.markdown("<br>", unsafe_allow_html=True)
        section("📊 P&L Aggregates")
        row("Total cost (EUR)",       "prod_cost + backfill_cost + training_cost + overhead_cost",
            "All cost lines combined.")
        row("Gross margin (EUR)",     "revenue − total_cost")
        row("Margin %",               "(gross_margin ÷ revenue) × 100")
        row("Break-even price",       "total_cost ÷ total_billable_hours",
            "Minimum EUR/hr you must charge to cover ALL costs.")
        row("Avg selling price",      "total_revenue ÷ total_billable_hours",
            "Blended rate across all blocks.")

        st.markdown("<br>", unsafe_allow_html=True)
        section("📋 Actual vs Budget")
        row("Revenue variance",       "actual_revenue − budget_revenue")
        row("Revenue variance %",     "(rev_variance ÷ budget_revenue) × 100")
        row("GM variance",            "actual_GM − budget_GM")
        row("YTD",                    "Sum of months where actuals have been entered")

        st.markdown("<br>", unsafe_allow_html=True)
        section("💱 FX / Currency")
        row("TRY cost → EUR",         "cost_TRY ÷ FX_rate",
            "FX = TRY per 1 EUR.")
        row("EUR revenue → TRY",      "revenue_EUR × FX_rate")
        row("Weighted avg FX",        "Σ(block_HC × block_FX) ÷ total_HC",
            "Used for backfill & training cost conversion when blocks have different FX overrides.")

        st.markdown("<br>", unsafe_allow_html=True)
        section("📈 FX Scenario Projection")
        row("Monthly FX (scenario)",  "start_FX + (end_FX − start_FX) × (month_idx ÷ 11)",
            "Linear interpolation: Jan = current rate, Dec = year-end scenario rate.")
        row("Scenario cost (EUR)",    "month_cost_TRY ÷ scenario_FX_month",
            "Higher TRY/EUR = your TRY costs are cheaper in EUR. Lower = more expensive.")
        row("Cost Δ vs budget",       "scenario_cost_EUR − budget_cost_EUR",
            "Negative = FX depreciation helps EUR cost. Positive = TRY strengthened.")
        row("FY scenario margin %",   "(Σ revenue − Σ scenario_cost) ÷ Σ revenue × 100",
            "Revenue stays fixed in EUR. Only cost base shifts with FX movement.")

st.caption("CC Budget Tool · Streamlit · openpyxl · plotly")
